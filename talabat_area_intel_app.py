from __future__ import annotations

import html
import io
import json
import math
import os
from urllib.parse import quote, quote_plus
import threading
import time
import uuid
from datetime import datetime

import folium
import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from folium.plugins import Fullscreen, HeatMap, MousePosition

from outbound_prioritization import (
    MODEL_HELP,
    add_priority_scores,
    build_brand_prioritization_table,
    format_for_dashboard,
)
from pin_validation import parse_scrape_pin_or_raise_value_error
from streamlit_location import (
    ensure_scrape_location,
    get_scrape_location,
    seed_city_preset_if_changed,
    set_scrape_location,
    sync_legacy_pin_mirror,
)
try:
    from batch_scrape_client import run_dual_area_scrape_via_api
except ImportError:
    # Deployment-safe fallback when an older module version is present.
    from batch_scrape_client import run_batch_scrape_via_api as run_dual_area_scrape_via_api
from geo_utils import haversine_km, haversine_series_km_from_pin
from google_map_tiles import (
    ensure_google_map_tile_sessions,
    google_2d_tile_url_template,
    google_maps_tile_attribution,
)
from supply_overlay import normalize_supply_overlay_df
from uae_cities import UAE_CITY_DISPLAY, UAE_CITY_PRESETS

DEFAULT_PIN = (25.2048, 55.2708)

# More grid points + deeper scroll = more listing URLs merged (slower; watch SCRAPER_WALL_CLOCK_SEC on the API host).
_DEFAULT_MAX_SAMPLE_POINTS = 6
_DEFAULT_SPACING_KM = 1.8
_DEFAULT_SCROLL_ROUNDS = 6
_DEFAULT_SCROLL_WAIT_MS = 500
_DEFAULT_CONCURRENCY = 3

_CITY_SLUGS = ["dubai", "sharjah", "abudhabi", "alain", "ajman"]

# Product defaults (no client toggles): full grid + cuisine sweep, keep all listing rows, request Places enrichment.
_SCRAPE_DEDUPE_BY_VENDOR_URL = True
_SCRAPE_HIGH_VOLUME = True
_SCRAPE_MAX_SAMPLE_POINTS = 6
# Default max time for /result polling when the payload does not set scrape_wall_clock_sec.
_SCRAPE_CLIENT_TIMEOUT_SEC = 1300
# POST /scrape only enqueues ({request_id, queued}); polling uses /result. Use (connect, read) timeouts because
# Streamlit Cloud → raw ``http://`` IP can stall on **connect** (read-only bump does not help).
_SCRAPE_POST_TIMEOUT_SEC = float(os.getenv("SCRAPER_API_POST_TIMEOUT_SEC", "600"))
# Max **read** seconds waiting for the first POST /scrape response body (tiny JSON for async enqueue). If the path
# stalls, fail fast and probe /result/{X-Request-ID}; raise cap (Secrets or env) for legacy sync /scrape.
_SCRAPE_ENQUEUE_READ_CAP_SEC = float(os.getenv("SCRAPER_API_ENQUEUE_READ_CAP_SEC", "240"))


def _scrape_poll_budget_sec_for_payload(payload: dict) -> float:
    """Upper bound for /result polling; must exceed API scrape_wall_clock_sec when the client sends it."""
    wall = int((payload or {}).get("scrape_wall_clock_sec") or 0)
    budget = float(_SCRAPE_CLIENT_TIMEOUT_SEC)
    if wall > 0:
        budget = max(budget, float(wall) + 180.0)
    try:
        raw_pb = str(st.secrets.get("SCRAPER_POLL_TOTAL_TIMEOUT_SEC", "")).strip()
        if raw_pb:
            budget = max(300.0, min(float(raw_pb), 10800.0))
    except Exception:
        pass
    return min(budget, 10800.0)

_SCRAPE_PROFILES: dict[str, dict] = {
    # Quick baseline in constrained hosting.
    "Fast": {
        "high_volume": False,
        "max_sample_points": 20,
        "scroll_rounds": 6,
        "scroll_wait_ms": 500,
        "google_places_enrich": True,
        "enrich": False,
    },
    # Better coverage with moderate runtime.
    "Balanced": {
        "high_volume": False,
        "max_sample_points": 20,
        "scroll_rounds": 6,
        "scroll_wait_ms": 500,
        "google_places_enrich": True,
        "enrich": False,
    },
    # Highest listing coverage on shared hosts; vendor Playwright enrich stays off unless API sets SCRAPER_VENDOR_PAGE_ENRICH.
    "Complete": {
        "high_volume": True,
        "max_sample_points": 20,
        "scroll_rounds": 6,
        "scroll_wait_ms": 500,
        "google_places_enrich": True,
        "enrich": False,
    },
    # Dedicated worker: many grid points + vendor detail pages (requires API env SCRAPER_VENDOR_PAGE_ENRICH=1).
    "Worker (vendor pages)": {
        "high_volume": True,
        "max_sample_points": 90,
        "scroll_rounds": 8,
        "scroll_wait_ms": 800,
        "google_places_enrich": True,
        "enrich": True,
        "scrape_wall_clock_sec": 2400,
    },
}
_DEFAULT_SCRAPE_PROFILE = "Complete"
_BUILD_STAMP = os.getenv("APP_BUILD_STAMP", "2026-04-23-executive-mode-80dbf21")


def inject_ui_theme() -> None:
    """Apply a cleaner, executive-style UI theme for Streamlit controls and sections."""
    st.markdown(
        """
<style>
:root {
  --bg: #f6f8fc;
  --card: #ffffff;
  --text: #0f172a;
  --muted: #475569;
  --line: #e2e8f0;
  --brand: #2563eb;
}
.stApp {
  background: linear-gradient(180deg, #f8fbff 0%, var(--bg) 280px, var(--bg) 100%);
}
.block-container {
  max-width: 1380px;
  padding-top: 1.2rem;
  padding-bottom: 2rem;
}
h1, h2, h3 {
  letter-spacing: -0.02em;
}
[data-testid="stSidebar"] {
  border-right: 1px solid var(--line);
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
  padding-top: 0.4rem;
}
div[data-testid="stMetric"] {
  background: var(--card);
  border: 1px solid var(--line);
  border-radius: 14px;
  padding: 0.65rem 0.85rem 0.55rem 0.85rem;
  box-shadow: 0 1px 2px rgba(2, 6, 23, 0.04);
}
div[data-testid="stAlert"] {
  border-radius: 12px;
}
div[data-testid="stDataFrame"] {
  border: 1px solid var(--line);
  border-radius: 12px;
  overflow: hidden;
  background: #fff;
}
button[kind="primary"] {
  border-radius: 999px !important;
  font-weight: 600 !important;
}
button[kind="secondary"] {
  border-radius: 999px !important;
}
.exec-pill {
  display: inline-block;
  padding: 0.22rem 0.65rem;
  border-radius: 999px;
  background: #dbeafe;
  color: #1d4ed8;
  font-size: 0.78rem;
  font-weight: 600;
  margin-left: 0.4rem;
}
.platform-pill {
  display: inline-block;
  padding: 0.2rem 0.58rem;
  border-radius: 999px;
  font-size: 0.74rem;
  font-weight: 700;
  margin-left: 0.38rem;
  border: 1px solid transparent;
}
.platform-pill.active {
  background: #fee2e2;
  color: #991b1b;
  border-color: #fecaca;
}
.platform-pill.soon {
  background: #f1f5f9;
  color: #64748b;
  border-color: #e2e8f0;
}
.step-card {
  border: 1px solid #e2e8f0;
  border-radius: 12px;
  background: #ffffff;
  padding: 0.7rem 0.85rem;
}
.step-card.done {
  border-color: #86efac;
  background: #f0fdf4;
}
.step-card.active {
  border-color: #93c5fd;
  background: #eff6ff;
}
</style>
        """,
        unsafe_allow_html=True,
    )


def init_state() -> None:
    # Hard-reset deprecated widget-bound keys that caused StreamlitAPIException in older builds.
    # Do this only once per browser session; popping on every rerun can clobber active pin widgets.
    if not st.session_state.get("_init_done"):
        st.session_state.pop("run_pin_lat__custom_pin_mode", None)
        st.session_state.pop("run_pin_lng__custom_pin_mode", None)
        st.session_state["_init_done"] = True
    ensure_scrape_location(
        default_lat=float(DEFAULT_PIN[0]),
        default_lng=float(DEFAULT_PIN[1]),
        default_label="Pin not set yet",
        migrate_from_legacy_keys=True,
    )
    sync_legacy_pin_mirror()
    st.session_state.setdefault("results_df", pd.DataFrame())
    st.session_state.setdefault("last_successful_results_df", pd.DataFrame())
    st.session_state.setdefault("last_run_done", False)
    st.session_state.setdefault("last_run_returned_zero", False)
    st.session_state.setdefault("results_fingerprint", None)
    st.session_state.setdefault("last_scrape_run_meta", {})
    st.session_state.setdefault("_last_successful_run_effective_pin", None)
    st.session_state.setdefault("last_geocode_provider", None)
    st.session_state.setdefault("last_geocode_label", None)
    st.session_state.setdefault("supply_overlay_df", None)
    st.session_state.setdefault("google_coverage_df", pd.DataFrame())
    loc0 = get_scrape_location()
    lat0 = float(loc0["lat"])
    lng0 = float(loc0["lng"])
    # Second pin defaults a few km away so A/B are not identical on first open.
    st.session_state.setdefault("dual_area_a_lat", lat0)
    st.session_state.setdefault("dual_area_a_lng", lng0)
    st.session_state.setdefault("dual_area_b_lat", lat0 + 0.018)
    st.session_state.setdefault("dual_area_b_lng", lng0)
    st.session_state.setdefault("dual_area_a_label", "")
    st.session_state.setdefault("dual_area_b_label", "")
    st.session_state.setdefault("dual_map_next_slot", "A")
    st.session_state.setdefault("dual_last_click_sig", "")
    st.session_state.setdefault("runpin_last_click_sig", "")
    st.session_state.setdefault("_last_successful_snapshot_pin", None)
    st.session_state.setdefault("_last_successful_dual_snapshot", None)


def _coords_tuple_close(a: tuple[float, ...] | None, b: tuple[float, ...] | None, tol: float = 1e-5) -> bool:
    if not a or not b or len(a) != len(b):
        return False
    return all(abs(float(a[i]) - float(b[i])) <= tol for i in range(len(a)))


def _bounds_for_radius(lat: float, lng: float, radius_km: float, pad: float = 1.15) -> tuple[list[float], list[float]]:
    """South-west and north-east corners so the map frames pin + search radius."""
    r = max(radius_km, 0.5) * pad
    d_lat = r / 110.574
    cos_lat = max(0.25, math.cos(math.radians(lat)))
    d_lng = r / (111.32 * cos_lat)
    return [lat - d_lat, lng - d_lng], [lat + d_lat, lng + d_lng]


def _default_zoom_for_radius_km(radius_km: float) -> int:
    """
    Stable Leaflet zoom for the search radius without ``fit_bounds``.
    """
    r = float(radius_km)
    if r <= 5.5:
        return 13
    return 12


def _add_supply_overlay_feature_group(fmap: folium.Map, supply_df: pd.DataFrame | None) -> None:
    supply = normalize_supply_overlay_df(supply_df)
    if supply is None or supply.empty:
        return
    supply_fg = folium.FeatureGroup(name="Supply / Kitchen Park")
    for _, s in supply.iterrows():
        la, ln = float(s["lat"]), float(s["lng"])
        lab = str(s.get("label") or "")[:120]
        folium.CircleMarker(
            location=[la, ln],
            radius=7,
            color="#EA580C",
            weight=2,
            fill=True,
            fill_color="#F97316",
            fill_opacity=0.85,
            tooltip=lab or "Supply",
            popup=folium.Popup(html.escape(lab or f"{la:.5f},{ln:.5f}"), max_width=220),
        ).add_to(supply_fg)
    supply_fg.add_to(fmap)


def _add_google_coverage_feature_group(fmap: folium.Map, coverage_df: pd.DataFrame | None) -> None:
    if coverage_df is None or not isinstance(coverage_df, pd.DataFrame) or coverage_df.empty:
        return
    if "lat" not in coverage_df.columns or "lng" not in coverage_df.columns:
        return
    fg = folium.FeatureGroup(name="Google-only coverage")
    for _, row in coverage_df.iterrows():
        try:
            la = float(row["lat"])
            ln = float(row["lng"])
        except (TypeError, ValueError):
            continue
        nm = html.escape(str(row.get("name") or "Google place")[:120])
        rating = row.get("rating")
        rt = f" · ⭐ {rating}" if rating not in (None, "") else ""
        popup = folium.Popup(f"<b>{nm}</b>{rt}", max_width=240)
        folium.CircleMarker(
            location=[la, ln],
            radius=6,
            color="#15803D",
            weight=2,
            fill=True,
            fill_color="#22C55E",
            fill_opacity=0.85,
            tooltip=f"{str(row.get('name') or 'Google place')[:80]}{rt}",
            popup=popup,
        ).add_to(fg)
    fg.add_to(fmap)


def _get_google_maps_api_key_for_basemap() -> str:
    try:
        secret = str(st.secrets.get("GOOGLE_MAPS_API_KEY", "")).strip()
        if secret:
            return secret
    except Exception:
        pass
    return (os.getenv("GOOGLE_MAPS_API_KEY") or "").strip()


def _render_google_maps_reference_panel(radius_km: float) -> None:
    """
    Google Maps as an **add-on** (embed + external links). Does not replace Folium: pin/scrape stay on the Search map.
    """
    loc = get_scrape_location()
    try:
        lat = float(loc["lat"])
        lng = float(loc["lng"])
    except (TypeError, ValueError, KeyError):
        return
    z = max(3, min(21, int(st.session_state.get("_pin_map_zoom_ui", 12))))
    key = _get_google_maps_api_key_for_basemap()
    embed_off = (os.getenv("STREAMLIT_GOOGLE_MAPS_EMBED", "1").strip().lower() in ("0", "false", "no", "off"))

    q_enc = quote_plus(f"{lat},{lng}")
    maps_open_url = f"https://www.google.com/maps/search/?api=1&query={q_enc}"
    directions_url = f"https://www.google.com/maps/dir/?api=1&destination={lat},{lng}"

    with st.expander("Google Maps (reference — same pin; does not replace the Search map)", expanded=False):
        st.caption(
            f"The **Interactive search map** above sets the Talabat scrape pin. This block is **Google Maps** only "
            f"for context (roads, labels, satellite). Radius context: **{radius_km:g} km**."
        )
        c1, c2 = st.columns(2)
        with c1:
            st.link_button("Open in Google Maps", maps_open_url, use_container_width=True)
        with c2:
            st.link_button("Directions to pin", directions_url, use_container_width=True)

        if key and not embed_off:
            mt = st.radio(
                "Embedded preview type",
                options=["roadmap", "satellite"],
                horizontal=True,
                index=0,
                key="google_embed_maptype_ref",
            )
            mt_param = "satellite" if mt == "satellite" else "roadmap"
            embed_src = (
                "https://www.google.com/maps/embed/v1/view?"
                f"key={quote(key, safe='')}&center={lat},{lng}&zoom={z}&maptype={mt_param}"
            )
            components.iframe(embed_src, height=420)
            st.caption(
                "Embedded via **Maps Embed API** (enable it for this key in Google Cloud). "
                "Restrict the key by HTTP referrer to your Streamlit URL."
            )
        elif not key:
            st.info(
                "Optional embedded preview: set **GOOGLE_MAPS_API_KEY** (same as Map Tiles) and enable **Maps Embed API**."
            )
        else:
            st.caption("Embedded preview is off (`STREAMLIT_GOOGLE_MAPS_EMBED=0`).")


def render_google_maps_pin(lat: float, lng: float, api_key: str, radius_km: float) -> str:
    """Interactive Google map: click/drag marker updates Streamlit bridge input."""
    safe_key = quote(str(api_key or "").strip(), safe="")
    radius_m = max(100.0, float(radius_km) * 1000.0)
    return f"""
<div id="gm-pin-map" style="height:420px;border:0;border-radius:10px;"></div>
<script>
  function _setUrlPin(lat, lng) {{
    const url = new URL(window.parent.location.href);
    const nextLat = lat.toFixed(6);
    const nextLng = lng.toFixed(6);
    if ((url.searchParams.get('pin_lat') || '') === nextLat && (url.searchParams.get('pin_lng') || '') === nextLng) return;
    url.searchParams.set('pin_lat', nextLat);
    url.searchParams.set('pin_lng', nextLng);
    url.searchParams.set('_pin_ts', String(Date.now()));
    window.parent.location.href = url.toString();
  }}

  function initGmPinMap() {{
    const center = {{ lat: {float(lat):.8f}, lng: {float(lng):.8f} }};
    const map = new google.maps.Map(document.getElementById('gm-pin-map'), {{
      center: center, zoom: 13, mapTypeId: 'roadmap', gestureHandling: 'greedy'
    }});
    const marker = new google.maps.Marker({{
      position: center, map: map, draggable: true,
      icon: 'http://maps.google.com/mapfiles/ms/icons/blue-dot.png'
    }});
    const circle = new google.maps.Circle({{
      map: map, center: center, radius: {radius_m:.1f},
      strokeColor: '#1D4ED8', strokeOpacity: 0.95, strokeWeight: 2,
      fillColor: '#2563EB', fillOpacity: 0.12
    }});
    map.addListener('click', (e) => {{
      marker.setPosition(e.latLng);
      circle.setCenter(e.latLng);
      _setUrlPin(e.latLng.lat(), e.latLng.lng());
    }});
    marker.addListener('dragend', (e) => {{
      circle.setCenter(e.latLng);
      _setUrlPin(e.latLng.lat(), e.latLng.lng());
    }});
  }}
</script>
<script async defer src="https://maps.googleapis.com/maps/api/js?key={safe_key}&callback=initGmPinMap"></script>
"""


def _configure_map_basemaps(fmap: folium.Map) -> str:
    """
    Add basemap tile layers: **Google** (Map Tiles API) when ``GOOGLE_MAPS_API_KEY`` is set
    and ``STREAMLIT_MAP_BASEMAP`` is not ``esri``; otherwise **Esri** (no Google key required).

    Returns ``\"google\"`` or ``\"esri\"`` for UI captions.
    """
    prefer = os.getenv("STREAMLIT_MAP_BASEMAP", "").strip().lower()
    key = _get_google_maps_api_key_for_basemap()
    if prefer != "esri" and key:
        cache = st.session_state.setdefault("_google_map_tile_session_cache", {})
        lang = os.getenv("STREAMLIT_GOOGLE_MAP_TILE_LANGUAGE", "en-US").strip() or "en-US"
        reg = os.getenv("STREAMLIT_GOOGLE_MAP_TILE_REGION", "AE").strip() or "AE"
        rm_sess, sh_sess = ensure_google_map_tile_sessions(key, cache, language=lang, region=reg)
        if rm_sess and sh_sess:
            g_attr = google_maps_tile_attribution()
            folium.TileLayer(
                tiles=google_2d_tile_url_template(key, rm_sess),
                attr=g_attr,
                name="Google roadmap",
                max_zoom=22,
            ).add_to(fmap)
            folium.TileLayer(
                tiles=google_2d_tile_url_template(key, sh_sess),
                attr=g_attr,
                name="Google satellite (labels)",
                max_zoom=22,
            ).add_to(fmap)
            return "google"
    _add_esri_basemaps(fmap)
    return "esri"


def _add_esri_basemaps(fmap: folium.Map) -> None:
    """Esri street + satellite; satellite is imagery-only until the reference overlay is on."""
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Street_Map/MapServer/tile/{z}/{y}/{x}",
        attr=(
            'Tiles © <a href="https://www.esri.com/">Esri</a> '
            "(HERE, Garmin, OpenStreetMap contributors, GIS user community)"
        ),
        name="Street map (English labels)",
        max_zoom=19,
    ).add_to(fmap)
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr='Tiles © <a href="https://www.esri.com/">Esri</a> (Earthstar Geographics, USDA, USGS, AeroGRID, IGN)',
        name="Satellite (photos only — no text on this layer)",
        max_zoom=19,
    ).add_to(fmap)
    # Raster imagery has no labels; this Esri reference layer adds place / boundary names (English-biased).
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}",
        attr='Labels © <a href="https://www.esri.com/">Esri</a>',
        name="Place names overlay (English · use with satellite)",
        overlay=True,
        control=True,
        opacity=0.92,
        show=True,
        max_zoom=19,
    ).add_to(fmap)


def _heal_run_pin_widgets_if_stale_default(widget_scope: str) -> None:
    """
    ``st.number_input`` keys keep their own session_state. If the user moved the pin via map/geocode but the
    widgets were never reset (e.g. never popped), they can still read the initial Dubai defaults — then the
    mandatory ``set_scrape_location`` after the widgets overwrites the real pin (often when clicking
    **Start Scraping**).
    """
    loc = get_scrape_location()
    lat_k = f"run_pin_lat_input__{widget_scope}"
    lng_k = f"run_pin_lng_input__{widget_scope}"
    try:
        w_lat = float(st.session_state.get(lat_k, loc["lat"]))
        w_lng = float(st.session_state.get(lng_k, loc["lng"]))
    except (TypeError, ValueError):
        st.session_state[lat_k] = float(loc["lat"])
        st.session_state[lng_k] = float(loc["lng"])
        return
    p_lat, p_lng = float(loc["lat"]), float(loc["lng"])
    # Widget still essentially on the app default, but authoritative pin is elsewhere.
    d_w_def = haversine_km(w_lat, w_lng, float(DEFAULT_PIN[0]), float(DEFAULT_PIN[1]))
    if d_w_def < 0.02 and haversine_km(w_lat, w_lng, p_lat, p_lng) > 0.05:
        st.session_state[lat_k] = p_lat
        st.session_state[lng_k] = p_lng


def _coalesce_run_pin_inputs_vs_authoritative(
    w_lat: float,
    w_lng: float,
    auth: dict,
) -> tuple[float, float]:
    """
    Second line of defence after ``_heal_*``: number_input can still return the template default while
    ``scrape_location`` already reflects a map/geocode pin — ``set_scrape_location(..., manual_form)`` would
    then clobber the real pin right before **Start Scraping**.
    """
    a_lat, a_lng = float(auth["lat"]), float(auth["lng"])
    d_w_tpl = haversine_km(w_lat, w_lng, float(DEFAULT_PIN[0]), float(DEFAULT_PIN[1]))
    if d_w_tpl >= 0.025:
        return w_lat, w_lng
    if haversine_km(w_lat, w_lng, a_lat, a_lng) <= 0.08:
        return w_lat, w_lng
    # User may legitimately be placing the pin in the same neighbourhood as the template centre.
    if haversine_km(a_lat, a_lng, float(DEFAULT_PIN[0]), float(DEFAULT_PIN[1])) <= 0.6:
        return w_lat, w_lng
    return a_lat, a_lng


def render_heatmap(
    df: pd.DataFrame,
    pin_lat: float,
    pin_lng: float,
    radius_km: float,
    *,
    supply_df: pd.DataFrame | None = None,
    google_coverage_df: pd.DataFrame | None = None,
) -> None:
    st.subheader("Restaurant Density Heatmap")
    view_df = df.dropna(subset=["lat", "lng"]).copy()
    if view_df.empty:
        st.info("No coordinates available for heatmap.")
        return

    fmap = folium.Map(
        location=[float(pin_lat), float(pin_lng)],
        tiles=None,
        zoom_start=12,
        zoom_control=True,
        control_scale=True,
    )
    basemap = _configure_map_basemaps(fmap)

    slack_km = float(radius_km) + 0.4
    d_computed = haversine_series_km_from_pin(
        float(pin_lat), float(pin_lng), view_df["lat"], view_df["lng"]
    )
    if "distance_km_from_pin" in view_df.columns:
        d_col = pd.to_numeric(view_df["distance_km_from_pin"], errors="coerce")
        d_eff = d_col.where(d_col.notna(), d_computed)
    else:
        d_eff = d_computed
    view_df = view_df.loc[d_eff <= slack_km].copy()

    heat_rows: list[list[float]] = []
    for _, row in view_df.iterrows():
        try:
            la = float(row["lat"])
            ln = float(row["lng"])
        except (TypeError, ValueError):
            continue
        heat_rows.append([la, ln])

    heat_source = "platform"
    if not heat_rows and google_coverage_df is not None and isinstance(google_coverage_df, pd.DataFrame):
        if "lat" in google_coverage_df.columns and "lng" in google_coverage_df.columns:
            gc = google_coverage_df.copy()
            la_s = pd.to_numeric(gc["lat"], errors="coerce")
            ln_s = pd.to_numeric(gc["lng"], errors="coerce")
            gc = gc.loc[la_s.notna() & ln_s.notna()].copy()
            if not gc.empty:
                d_gc = haversine_series_km_from_pin(
                    float(pin_lat), float(pin_lng), gc["lat"], gc["lng"]
                )
                gc = gc.loc[d_gc <= slack_km].copy()
                for _, row in gc.iterrows():
                    try:
                        heat_rows.append([float(row["lat"]), float(row["lng"])])
                    except (TypeError, ValueError):
                        continue
                if heat_rows:
                    heat_source = "google_coverage"

    heat_layer_name = (
        "Platform listing density"
        if heat_source == "platform"
        else "Google coverage density (no platform coords in radius)"
    )
    if heat_rows:
        heat_fg = folium.FeatureGroup(name=heat_layer_name, overlay=True, control=True)
        HeatMap(
            heat_rows,
            min_opacity=0.4,
            max_zoom=18,
            radius=22,
            blur=14,
            gradient={0.4: "#2563EB", 0.6: "#7C3AED", 0.8: "#F59E0B", 0.96: "#EF4444"},
        ).add_to(heat_fg)
        heat_fg.add_to(fmap)
    else:
        st.warning(
            "No in-radius points with coordinates to plot. "
            "If the table has rows but no **lat**/**lng**, enable enrichment or widen radius; "
            "or use **Google-only coverage** when the platform scrape is empty."
        )

    folium.Circle(
        location=[float(pin_lat), float(pin_lng)],
        radius=float(radius_km) * 1000.0,
        color="#1D4ED8",
        weight=2,
        fill=True,
        fill_color="#2563EB",
        fill_opacity=0.08,
        tooltip=f"Scrape radius: {radius_km:g} km",
    ).add_to(fmap)
    folium.Marker(
        location=[float(pin_lat), float(pin_lng)],
        tooltip="Search pin",
        icon=folium.Icon(color="blue"),
    ).add_to(fmap)

    _add_supply_overlay_feature_group(fmap, supply_df)
    _add_google_coverage_feature_group(fmap, google_coverage_df)

    Fullscreen(position="topright", title="Fullscreen", title_cancel="Exit Full Screen").add_to(fmap)
    folium.LayerControl(position="topright", collapsed=False).add_to(fmap)

    lats = [float(pin_lat)] + [r[0] for r in heat_rows]
    lngs = [float(pin_lng)] + [r[1] for r in heat_rows]
    supply_bounds = normalize_supply_overlay_df(supply_df)
    if supply_bounds is not None and not supply_bounds.empty:
        lats.extend(supply_bounds["lat"].astype(float).tolist())
        lngs.extend(supply_bounds["lng"].astype(float).tolist())
    if google_coverage_df is not None and isinstance(google_coverage_df, pd.DataFrame) and not google_coverage_df.empty:
        if "lat" in google_coverage_df.columns and "lng" in google_coverage_df.columns:
            lats.extend(pd.to_numeric(google_coverage_df["lat"], errors="coerce").dropna().astype(float).tolist())
            lngs.extend(pd.to_numeric(google_coverage_df["lng"], errors="coerce").dropna().astype(float).tolist())
    pad_lat = max(0.002, (max(lats) - min(lats)) * 0.08 + 0.001)
    pad_lng = max(0.002, (max(lngs) - min(lngs)) * 0.08 + 0.001)
    sw = [min(lats) - pad_lat, min(lngs) - pad_lng]
    ne = [max(lats) + pad_lat, max(lngs) + pad_lng]
    fmap.fit_bounds([sw, ne], padding=(28, 28), max_zoom=16)

    if basemap == "google":
        st.caption(
            "Same **Google** basemaps as the pin map. "
            + (
                "**Heat** uses Google coverage points (Talabat had no coordinates in this view)."
                if heat_source == "google_coverage"
                else "**Heat** = **platform listing density** from this scrape (single aggregator)."
            )
        )
    else:
        st.caption(
            "Same English-first basemaps as the pin map. **Street** / **Satellite** + **Place names overlay** in the layer control. "
            + (
                "**Heat** uses Google coverage points (Talabat had no coordinates in this view)."
                if heat_source == "google_coverage"
                else "**Heat** = **platform listing density** from this scrape (single aggregator)."
            )
        )
    # Full HTML via ``components.html`` matches Folium's notebook/static path and reliably loads Leaflet + plugins.
    fig = folium.Figure().add_child(fmap)
    components.html(fig.render(), height=535, width=None, scrolling=False)


def get_frontend_api_key() -> str:
    try:
        secret = str(st.secrets.get("SCRAPER_API_KEY", "")).strip()
        if secret:
            return secret
    except Exception:
        pass
    return os.getenv("SCRAPER_API_KEY", "").strip()


def render_outbound_prioritization_dashboard(df: pd.DataFrame) -> None:
    """Brand-level outbound view: cuisine, order/store proxies, weighted priority score."""
    st.subheader("Outbound prioritization (brand view)")
    st.caption(
        "KitchenPark-style lens: rank brands for acquisition outreach using cuisine fit, ratings, reviews, "
        "delivery fee, and footprint. Talabat does not expose true **last-7-days** orders in this scrape — "
        "the order column is a **platform proxy** when available."
    )
    brand_raw = build_brand_prioritization_table(df)
    if brand_raw is None or brand_raw.empty:
        st.info("Not enough data to build a brand-level view.")
        return

    with st.expander("Prioritization model (weights & definitions)", expanded=False):
        st.markdown(MODEL_HELP)
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            w_r = st.slider("Rating", 0.0, 1.0, 0.28, 0.01, help="Talabat / Google rating blend per brand")
        with c2:
            w_rev = st.slider("Reviews", 0.0, 1.0, 0.22, 0.01, help="Sum of review counts across sampled branches")
        with c3:
            w_ord = st.slider("Order proxy", 0.0, 1.0, 0.22, 0.01, help="Talabat estimated_orders when present")
        with c4:
            w_del = st.slider("Delivery fee", 0.0, 1.0, 0.14, 0.01, help="Lower median fee → higher score")
        with c5:
            w_scale = st.slider("Store footprint", 0.0, 1.0, 0.14, 0.01, help="More stores in sample → higher reach")

    scored = add_priority_scores(
        brand_raw,
        w_rating=w_r,
        w_reviews=w_rev,
        w_orders=w_ord,
        w_delivery=w_del,
        w_scale=w_scale,
    )
    view = format_for_dashboard(scored)
    st.dataframe(view, width="stretch", height=380)

    top_n = view.head(18).copy()
    if not top_n.empty and "Outbound_priority" in top_n.columns:
        chart_df = top_n.set_index("Brand")["Outbound_priority"].sort_values(ascending=True)
        st.caption("Top brands by composite **Outbound_priority** (this scrape only).")
        st.bar_chart(chart_df)

    st.download_button(
        "Download brand prioritization CSV",
        data=view.to_csv(index=False).encode("utf-8"),
        file_name="talabat_outbound_brand_priorities.csv",
        mime="text/csv",
        key="dl_brand_priority",
    )


def _pick_numeric_series(df: pd.DataFrame, candidates: list[str]) -> pd.Series:
    for col in candidates:
        if col in df.columns:
            return pd.to_numeric(df[col], errors="coerce")
    return pd.Series(dtype=float)


def _pick_rating_series(df: pd.DataFrame) -> pd.Series:
    return _pick_numeric_series(df, ["rating_effective", "rating", "google_rating"])


def _pick_area_group_series(df: pd.DataFrame) -> pd.Series:
    for col in ("batch_location_label", "scrape_target_label", "dual_area"):
        if col in df.columns:
            s = df[col].astype(str).str.strip()
            s = s.mask(s == "", "Unlabeled area")
            return s
    return pd.Series(["Current run"] * len(df), index=df.index)


def render_executive_mode(df: pd.DataFrame, meta: dict) -> None:
    """Leadership-ready summary: KPI strip, run confidence, A/B delta, and area opportunity ranking."""
    st.subheader("Executive Mode")
    st.caption("GM view: headline KPIs, run confidence, dual-area delta, and top opportunity areas.")

    rating_s = _pick_rating_series(df)
    orders_day_s = _pick_numeric_series(df, ["estimated_orders_per_day", "estimated_orders"])
    orders_week_s = _pick_numeric_series(df, ["estimated_orders_per_week"])
    brand_key_s = (
        df["brand_id"].astype(str).str.strip()
        if "brand_id" in df.columns
        else df.get("restaurant_name", pd.Series([""] * len(df))).astype(str).str.strip().str.lower()
    )
    unique_brands = int((brand_key_s.replace("", pd.NA).dropna()).nunique()) if len(brand_key_s) else 0

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Rows", f"{len(df):,}")
    k2.metric("Unique brands", f"{unique_brands:,}")
    avg_rating = float(rating_s.dropna().mean()) if not rating_s.dropna().empty else 0.0
    k3.metric("Avg rating", f"{avg_rating:.2f}" if avg_rating > 0 else "—")
    weekly_proxy = float(orders_week_s.dropna().sum()) if not orders_week_s.dropna().empty else 0.0
    if weekly_proxy <= 0:
        weekly_proxy = float(orders_day_s.dropna().sum() * 7.0) if not orders_day_s.dropna().empty else 0.0
    k4.metric("Weekly orders proxy", f"{int(round(weekly_proxy)):,}" if weekly_proxy > 0 else "—")

    legal_cov = 0.0
    if "legal_name" in df.columns and len(df):
        legal_cov = float((df["legal_name"].astype(str).str.strip() != "").mean())
    phone_cov = 0.0
    if "contact_phone" in df.columns and len(df):
        phone_cov = float((df["contact_phone"].astype(str).str.strip() != "").mean())
    elif "phone" in df.columns and len(df):
        phone_cov = float((df["phone"].astype(str).str.strip() != "").mean())
    rating_cov = float(rating_s.notna().mean()) if len(df) else 0.0
    geo_cov = 0.0
    if {"lat", "lng"}.issubset(df.columns) and len(df):
        lat_ok = pd.to_numeric(df["lat"], errors="coerce").notna()
        lng_ok = pd.to_numeric(df["lng"], errors="coerce").notna()
        geo_cov = float((lat_ok & lng_ok).mean())
    scale_score = min(float(len(df)) / 250.0, 1.0)
    meta_score = 1.0 if meta.get("request_id") else 0.0
    confidence = int(round(100.0 * (0.20 * legal_cov + 0.15 * phone_cov + 0.15 * rating_cov + 0.15 * geo_cov + 0.25 * scale_score + 0.10 * meta_score)))
    conf_txt = "High" if confidence >= 75 else ("Medium" if confidence >= 50 else "Low")
    st.info(
        f"Run confidence: **{confidence}/100 ({conf_txt})** · "
        f"legal `{legal_cov*100:.0f}%` · phone `{phone_cov*100:.0f}%` · rating `{rating_cov*100:.0f}%` · geo `{geo_cov*100:.0f}%`."
    )

    if "dual_area" in df.columns and {"A", "B"}.issubset(set(df["dual_area"].astype(str).str.upper().str.strip())):
        comp = df.copy()
        comp["dual_area"] = comp["dual_area"].astype(str).str.upper().str.strip()
        ab = comp[comp["dual_area"].isin(["A", "B"])].copy()
        ab["__rating"] = _pick_rating_series(ab)
        ab["__orders_day"] = _pick_numeric_series(ab, ["estimated_orders_per_day", "estimated_orders"])
        agg_dict = {
            "rows": ("dual_area", "size"),
            "avg_rating": ("__rating", "mean"),
            "orders_day": ("__orders_day", "sum"),
        }
        if "brand_id" in ab.columns:
            agg_dict["brands"] = ("brand_id", lambda s: s.astype(str).str.strip().replace("", pd.NA).dropna().nunique())
        else:
            agg_dict["brands"] = ("restaurant_name", "nunique")
        agg = ab.groupby("dual_area", dropna=False).agg(**agg_dict)
        if {"A", "B"}.issubset(set(agg.index)):
            a = agg.loc["A"]
            b = agg.loc["B"]
            st.markdown("**Area A vs B (executive delta)**")
            d1, d2, d3, d4 = st.columns(4)
            d1.metric("Rows (A)", f"{int(a['rows']):,}", delta=f"{int(a['rows'] - b['rows']):+d} vs B")
            d2.metric("Brands (A)", f"{int(a['brands']):,}", delta=f"{int(a['brands'] - b['brands']):+d} vs B")
            if pd.notna(a["avg_rating"]) and pd.notna(b["avg_rating"]):
                d3.metric("Avg rating (A)", f"{float(a['avg_rating']):.2f}", delta=f"{float(a['avg_rating'] - b['avg_rating']):+.2f} vs B")
            else:
                d3.metric("Avg rating (A)", "—")
            if pd.notna(a["orders_day"]) and pd.notna(b["orders_day"]):
                d4.metric("Orders/day proxy (A)", f"{int(round(float(a['orders_day']))):,}", delta=f"{int(round(float(a['orders_day'] - b['orders_day']))):+d} vs B")
            else:
                d4.metric("Orders/day proxy (A)", "—")

    area_s = _pick_area_group_series(df)
    work = df.copy()
    work["__area"] = area_s
    work["__rating"] = _pick_rating_series(work).fillna(0.0)
    work["__orders"] = _pick_numeric_series(work, ["estimated_orders_per_day", "estimated_orders"]).fillna(0.0)
    if "restaurant_name" not in work.columns:
        work["restaurant_name"] = ""
    grp = work.groupby("__area", dropna=False).agg(
        rows=("__area", "size"),
        avg_rating=("__rating", "mean"),
        orders_proxy=("__orders", "sum"),
        restaurants=("restaurant_name", "nunique"),
    ).reset_index()
    if len(grp) > 0:
        def _norm(s: pd.Series) -> pd.Series:
            s = pd.to_numeric(s, errors="coerce").fillna(0.0)
            lo, hi = float(s.min()), float(s.max())
            if hi <= lo:
                return pd.Series([0.5] * len(s), index=s.index)
            return (s - lo) / (hi - lo)

        grp["score_orders"] = _norm(grp["orders_proxy"])
        grp["score_scale"] = _norm(grp["restaurants"])
        grp["score_rating_gap"] = _norm(4.3 - grp["avg_rating"])
        grp["opportunity_score"] = (0.45 * grp["score_orders"] + 0.30 * grp["score_scale"] + 0.25 * grp["score_rating_gap"]) * 100.0
        grp = grp.sort_values(["opportunity_score", "orders_proxy"], ascending=[False, False]).reset_index(drop=True)
        top = grp.head(10).copy()
        top["opportunity_score"] = top["opportunity_score"].round(1)
        top["avg_rating"] = top["avg_rating"].round(2)
        st.markdown("**Top opportunity areas**")
        st.dataframe(
            top.rename(columns={"__area": "Area"})[["Area", "opportunity_score", "orders_proxy", "restaurants", "avg_rating", "rows"]],
            width="stretch",
            height=260,
        )
        if not top.empty:
            lead = top.iloc[0]
            st.success(
                f"Primary recommendation: prioritize **{lead['__area']}** first "
                f"(score {float(lead['opportunity_score']):.1f}, orders proxy {int(round(float(lead['orders_proxy']))):,})."
            )


def get_api_base_url() -> str:
    try:
        secret_url = str(st.secrets.get("API_BASE_URL", "")).strip()
        if secret_url:
            return secret_url.rstrip("/")
    except Exception:
        pass
    return os.getenv("API_BASE_URL", "https://maisam21-lab-talabat-area-intel.onrender.com").strip().rstrip("/")


_PLACEHOLDER_CELL_LOWER = frozenset(
    {
        "",
        "unknown",
        "n/a",
        "na",
        "none",
        "null",
        "-",
        "—",
        "nan",
        "undefined",
        "n.a.",
        "(null)",
    }
)


def _meaningful_value_mask(s: pd.Series) -> pd.Series:
    """True where a cell counts as real data (not blank / common sentinel strings)."""
    if pd.api.types.is_numeric_dtype(s):
        return s.notna()
    if pd.api.types.is_bool_dtype(s):
        return s.notna()
    if pd.api.types.is_datetime64_any_dtype(s):
        return s.notna()
    t = s.astype(str).str.strip()
    low = t.str.lower()
    return s.notna() & ~low.isin(_PLACEHOLDER_CELL_LOWER)


def polish_dataframe_display_noise(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize noisy empty sentinels to blank cells for table / CSV / Excel."""
    if df is None or df.empty:
        return df
    out = df.copy()

    def _clean_cell(val: object) -> object:
        if val is None:
            return ""
        if isinstance(val, float) and pd.isna(val):
            return ""
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return val
        s = str(val).strip()
        if s.lower() in _PLACEHOLDER_CELL_LOWER:
            return ""
        return s

    for c in out.columns:
        ser = out[c]
        if pd.api.types.is_numeric_dtype(ser) or pd.api.types.is_bool_dtype(ser):
            continue
        if pd.api.types.is_datetime64_any_dtype(ser):
            continue
        if hasattr(ser.dtype, "categories"):
            out[c] = ser.astype(str).map(_clean_cell)
            continue
        if not (pd.api.types.is_object_dtype(ser) or pd.api.types.is_string_dtype(ser)):
            continue
        out[c] = ser.map(_clean_cell)
    return out


def _friendly_api_error(response: requests.Response) -> str:
    rid = (response.headers.get("X-Request-ID") or "").strip()
    code = int(response.status_code)
    ctype = (response.headers.get("Content-Type") or "").lower()
    if "application/json" in ctype:
        try:
            data = response.json()
        except ValueError:
            data = {}
        detail = str(data.get("error") or data.get("detail") or response.reason or "Request failed")
        rid = str(data.get("request_id") or rid).strip()
    else:
        detail = "Upstream gateway returned a non-JSON error page before the API could send structured JSON."
    hint = {
        502: "Try a smaller radius or check API logs for this request id.",
        504: "Scrape timed out. Raise server timeout limits or check API logs.",
        500: "Internal scrape failure. Check backend logs with request id.",
    }.get(code, "Check API logs with request id.")
    rid_txt = f" Request ID: {rid}." if rid else ""
    return f"{code} {response.reason}: {detail}. {hint}{rid_txt}"


def compact_output_df(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """
    Remove all-empty columns from user-facing outputs.

    Keeps only core geometry/identity columns by default, then appends any
    additional columns that have at least one non-empty value.
    """
    if df is None or df.empty:
        return pd.DataFrame(), []
    core = [
        "restaurant_name",
        "brand_display_name",
        "restaurant_url",
        "legal_name",
        "legal_name_final",
        "legal_name_candidate",
        "legal_name_source",
        "contact_phone",
        "contact_final",
        "contact_source_final",
        "outside_talabat_contact_mapped",
        "contact_source",
        "vendor_email",
        "vendor_website",
        "vendor_social",
        "required_fields_ready",
        "source_url",
        "last_verified_at",
        "confidence_score",
        "tax_or_license_hint",
        "just_landed",
        "just_landed_date",
        "rating",
        "google_rating",
        "rating_effective",
        "rating_final",
        "estimated_orders",
        "estimated_orders_per_day",
        "estimated_orders_per_week",
        "orders_final",
        "cuisine_final",
        "lat",
        "lng",
        "distance_km_from_pin",
    ]
    keep: list[str] = [c for c in core if c in df.columns]
    removed: list[str] = []

    for c in df.columns:
        if c in keep:
            continue
        s = df[c]
        if bool(_meaningful_value_mask(s).any()):
            keep.append(c)
        else:
            removed.append(c)
    return df.loc[:, keep].copy(), removed


def is_google_coverage_only_results(df: pd.DataFrame) -> bool:
    """True when every row is the Google coverage fallback (Talabat scrape returned no restaurants)."""
    if df is None or df.empty or "platform" not in df.columns:
        return False
    v = df["platform"].astype(str).str.strip().str.lower()
    return bool(len(v) and (v == "google coverage").all())


def ensure_just_landed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep Just Landed signal visible in UI/exports with stable defaults."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "just_landed" not in out.columns:
        out["just_landed"] = "false"
    jl = out["just_landed"].astype(str).str.strip().str.lower()
    out["just_landed"] = jl.map({"yes": "true", "true": "true"}).fillna("false")
    if "just_landed_date" not in out.columns:
        out["just_landed_date"] = ""
    out["just_landed_date"] = out["just_landed_date"].fillna("").astype(str).str.strip()
    return out


def build_sanity_check_report(df: pd.DataFrame, radius_km: float) -> dict[str, float | int | str]:
    """Compute lightweight quality sanity metrics for stakeholder validation."""
    if df is None or df.empty:
        return {
            "rows_total": 0,
            "unique_brands": 0,
            "unique_brand_ratio": 0.0,
            "rows_with_contact": 0,
            "contact_coverage_pct": 0.0,
            "rows_with_legal_name": 0,
            "legal_name_coverage_pct": 0.0,
            "unique_cuisines": 0,
            "top_cuisine_share_pct": 0.0,
            "status": "warn",
            "note": "No rows returned.",
        }

    total = int(len(df))
    bsrc = df.get("brand_display_name", df.get("restaurant_name", pd.Series([""] * total)))
    brands = bsrc.astype(str).str.strip().replace("", pd.NA).dropna()
    unique_brands = int(brands.nunique())
    unique_brand_ratio = (unique_brands / total) if total else 0.0

    has_contact = (
        df.get("contact_phone", pd.Series([""] * total)).astype(str).str.strip().replace("", pd.NA).notna()
        | df.get("vendor_email", pd.Series([""] * total)).astype(str).str.strip().replace("", pd.NA).notna()
        | df.get("vendor_website", pd.Series([""] * total)).astype(str).str.strip().replace("", pd.NA).notna()
    )
    rows_with_contact = int(has_contact.sum())
    contact_coverage_pct = (rows_with_contact * 100.0 / total) if total else 0.0

    legal = df.get("legal_name", pd.Series([""] * total)).astype(str).str.strip().replace("", pd.NA).notna()
    rows_with_legal_name = int(legal.sum())
    legal_name_coverage_pct = (rows_with_legal_name * 100.0 / total) if total else 0.0

    cuisines_raw = df.get("cuisines", pd.Series([""] * total)).fillna("").astype(str).str.strip()
    cuisine_tokens = cuisines_raw.str.split(",").explode().astype(str).str.strip()
    cuisine_tokens = cuisine_tokens[cuisine_tokens != ""]
    unique_cuisines = int(cuisine_tokens.nunique()) if not cuisine_tokens.empty else 0
    if cuisine_tokens.empty:
        top_cuisine_share_pct = 0.0
    else:
        counts = cuisine_tokens.value_counts()
        top_cuisine_share_pct = float((counts.iloc[0] * 100.0) / max(1, int(counts.sum())))

    # Simple operational guardrails; tuned to catch obvious anomalies, not strict rejection criteria.
    min_brands_expected = 50 if float(radius_km) >= 10.0 else 20
    status = "ok"
    notes: list[str] = []
    if unique_brands < min_brands_expected:
        status = "warn"
        notes.append(f"Low unique brand count ({unique_brands} < {min_brands_expected} expected baseline).")
    if unique_brand_ratio < 0.08:
        status = "warn"
        notes.append(f"Low unique-brand ratio ({unique_brand_ratio:.3f}).")
    if unique_cuisines <= 5 or top_cuisine_share_pct >= 55.0:
        status = "warn"
        notes.append("Cuisine diversity looks weak; verify parsing/mapping.")
    if contact_coverage_pct < 8.0:
        status = "warn"
        notes.append("Contact coverage is low; legal/contact enrichment still incomplete.")

    return {
        "rows_total": total,
        "unique_brands": unique_brands,
        "unique_brand_ratio": round(unique_brand_ratio, 4),
        "rows_with_contact": rows_with_contact,
        "contact_coverage_pct": round(contact_coverage_pct, 2),
        "rows_with_legal_name": rows_with_legal_name,
        "legal_name_coverage_pct": round(legal_name_coverage_pct, 2),
        "unique_cuisines": unique_cuisines,
        "top_cuisine_share_pct": round(top_cuisine_share_pct, 2),
        "status": status,
        "note": " ".join(notes) if notes else "Sanity checks look reasonable for this run.",
    }


def build_excel_export_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "platform" not in out.columns:
        out["platform"] = "Talabat"
    if "orders_week_estimate" not in out.columns:
        if "estimated_orders_per_week" in out.columns:
            out["orders_week_estimate"] = out["estimated_orders_per_week"]
        elif "estimated_orders_per_day" in out.columns:
            out["orders_week_estimate"] = pd.to_numeric(out["estimated_orders_per_day"], errors="coerce") * 7.0
        else:
            out["orders_week_estimate"] = ""
    if "area_label" not in out.columns:
        if "scrape_target_label" in out.columns:
            out["area_label"] = out["scrape_target_label"]
        elif "batch_location_label" in out.columns:
            out["area_label"] = out["batch_location_label"]
        else:
            out["area_label"] = ""
    out["scrape_date"] = datetime.utcnow().strftime("%Y-%m-%d")
    # Legal + contact first (vendor page / Places enrichment); identity URLs next.
    export_cols = [
        "restaurant_name",
        "brand_display_name",
        "legal_name",
        "legal_name_final",
        "legal_name_candidate",
        "legal_name_source",
        "contact_phone",
        "contact_final",
        "contact_source_final",
        "outside_talabat_contact_mapped",
        "contact_source",
        "vendor_email",
        "vendor_website",
        "vendor_social",
        "required_fields_ready",
        "source_url",
        "last_verified_at",
        "confidence_score",
        "tax_or_license_hint",
        "restaurant_url",
        "brand_id",
        "branch_sku",
        "cuisines",
        "cuisine_final",
        "rating_final",
        "orders_week_estimate",
        "orders_final",
        "platform",
        "status",
        "just_landed",
        "just_landed_date",
        "lat",
        "lng",
        "distance_km_from_pin",
        "area_label",
        "scrape_date",
    ]
    for c in export_cols:
        if c not in out.columns:
            out[c] = ""
    return out.loc[:, export_cols].rename(
        columns={
            "restaurant_name": "restaurant",
            "brand_display_name": "brand",
            "legal_name": "legal_entity_name",
            "legal_name_final": "legal_name_final",
            "contact_phone": "phone",
            "contact_final": "contact_final",
            "vendor_email": "email",
            "vendor_website": "website",
            "vendor_social": "social_links",
            "tax_or_license_hint": "tax_or_license",
            "cuisines": "cuisine",
            "cuisine_final": "cuisine_final",
            "distance_km_from_pin": "distance_from_pin_km",
            "restaurant_url": "listing_url",
        }
    )


def _apply_area_intel_excel_formatting(ws, df: pd.DataFrame) -> None:
    """Header styling, freeze pane, filters, widths, and sensible number formats."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1 or not len(df.columns):
        return
    header_fill = PatternFill(start_color="FFEFF6FF", end_color="FFEFF6FF", fill_type="solid")
    bold = Font(bold=True, color="FF1E3A8A")
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    lower_names = [str(c).lower() for c in df.columns]
    fmt_for_col: dict[int, str] = {}
    for idx, name in enumerate(lower_names, start=1):
        if name in ("lat", "lng"):
            fmt_for_col[idx] = "0.000000"
        elif "distance" in name and "km" in name:
            fmt_for_col[idx] = "0.00"
        elif "orders" in name or "estimate" in name:
            fmt_for_col[idx] = "0"

    for col_idx in range(1, len(df.columns) + 1):
        letter = get_column_letter(col_idx)
        hdr = str(df.columns[col_idx - 1])
        max_len = min(max(len(hdr), 12), 80)
        sample_rows = min(ws.max_row, 600)
        for row in range(2, sample_rows + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None and str(val).strip() != "":
                max_len = min(max(max_len, len(str(val))), 80)
        ws.column_dimensions[letter].width = min(max(max_len + 1.5, 11), 52)
        if col_idx in fmt_for_col:
            fc = fmt_for_col[col_idx]
            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=col_idx)
                if c.value is not None and c.value != "":
                    try:
                        if isinstance(c.value, (int, float)) or str(c.value).strip() != "":
                            c.number_format = fc
                    except Exception:
                        pass


def dataframe_to_excel_bytes(df: pd.DataFrame, *, summary: dict[str, str] | None = None) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Area Intel")
        ws_data = writer.book["Area Intel"]
        _apply_area_intel_excel_formatting(ws_data, df)
        if summary:
            from openpyxl.styles import Alignment, Font

            meta_ws = writer.book.create_sheet("Run summary", 0)
            meta_ws["A1"], meta_ws["B1"] = "Field", "Value"
            meta_ws["A1"].font = Font(bold=True)
            meta_ws["B1"].font = Font(bold=True)
            row = 2
            for k, v in summary.items():
                meta_ws.cell(row=row, column=1, value=k)
                meta_ws.cell(row=row, column=2, value=v)
                meta_ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
            meta_ws.column_dimensions["A"].width = 28
            meta_ws.column_dimensions["B"].width = 72
    return bio.getvalue()


def main() -> None:
    st.set_page_config(page_title="Area Intel | Kitchen Park", layout="wide")
    inject_ui_theme()
    init_state()

    st.markdown(
        "## Area Intel "
        "<span class='platform-pill active'>Talabat</span>"
        "<span class='platform-pill soon'>Deliveroo · Coming soon</span>"
        "<span class='platform-pill soon'>Careem · Coming soon</span>",
        unsafe_allow_html=True,
    )
    st.caption("Kitchen Park · Market Intelligence")
    st.warning(f"Build: `{_BUILD_STAMP}`")

    with st.sidebar:
        st.header("Scrape Controls")
        api_base_url = get_api_base_url()
        api_key = get_frontend_api_key()
        headers = {"X-API-Key": api_key} if api_key else {}
        show_advanced = False
        area_mode = "custom"
        is_city_mode = False
        city_key = "dubai"

        selected_profile_name = st.selectbox(
            "Scrape profile",
            options=list(_SCRAPE_PROFILES.keys()),
            index=list(_SCRAPE_PROFILES.keys()).index("Complete"),
            help=(
                "**Worker (vendor pages)** = sends enrich=true for vendor detail pages; use when the API host has "
                "SCRAPER_VENDOR_PAGE_ENRICH=1 (e.g. Hetzner worker)."
            ),
            key="sidebar_scrape_profile",
        )
        st.caption(f"API: `{api_base_url}` · Profiles: {', '.join(_SCRAPE_PROFILES.keys())}")

        st.markdown("**1) Location**")
        city_key = st.selectbox(
            "City preset",
            _CITY_SLUGS,
            format_func=lambda k: UAE_CITY_DISPLAY[k],
            index=0,
        )
        city_lat, city_lng, _ = UAE_CITY_PRESETS[city_key]
        if st.button("Use city preset pin", width="stretch"):
            seed_city_preset_if_changed(
                city_key,
                float(city_lat),
                float(city_lng),
                UAE_CITY_DISPLAY[city_key],
            )
            sync_legacy_pin_mirror()
            st.rerun()

        target_area_label = ""
        listing_status_mode = "live"
        new_on_platform_only = st.checkbox(
            "Just Landed only",
            value=False,
            help="When enabled, only vendors marked as newly launched are included.",
        )
        include_google_coverage = True
        radius_pick = st.radio("2) Radius", options=[5, 10], horizontal=True, index=1)
        radius_km = float(radius_pick)
        target_area_label = st.text_input(
            "3) Area label (for exports)",
            value="",
            help="Stored as area_label in exports.",
        )

        geocode_query = st.text_input(
            "Address search (UAE)",
            value="",
            help="Enter a place name or street (e.g. Dubai Marina). The API rejects an empty search.",
        )
        _geocode_has_query = bool((geocode_query or "").strip())
        geocode_btn = st.button(
            "Set pin from search",
            width="stretch",
            disabled=not _geocode_has_query,
            help=("Type an address above first.") if not _geocode_has_query else None,
        )

        if geocode_btn:
            q = (geocode_query or "").strip()
            if not q:
                st.warning("Enter an address or area in **Address search (UAE)** before clicking **Set pin from search**.")
            else:
                try:
                    g_response = requests.post(
                        f"{api_base_url.rstrip('/')}/geocode",
                        json={"query": q},
                        headers=headers,
                        timeout=30,
                    )
                    if not g_response.ok:
                        try:
                            body = g_response.json()
                            msg = body.get("error") or body.get("detail") or g_response.reason
                        except Exception:
                            msg = g_response.text or g_response.reason
                        st.error(f"Geocode failed ({g_response.status_code}): {msg}")
                    else:
                        payload = g_response.json()
                        result = payload.get("result")
                        if payload.get("ok") and result:
                            set_scrape_location(
                                float(result["lat"]),
                                float(result["lng"]),
                                str(result.get("formatted_address") or q).strip(),
                                "geocode",
                            )
                            sync_legacy_pin_mirror()
                            provider = payload.get("provider", "unknown")
                            st.session_state["last_geocode_provider"] = str(provider)
                            st.session_state["last_geocode_label"] = str(get_scrape_location().get("label") or "")
                            st.success(f"Pin set from search ({provider}): {get_scrape_location()['label']}")
                            if payload.get("note"):
                                st.info(str(payload["note"]))
                        else:
                            hint = payload.get("hint")
                            if hint:
                                st.warning(hint)
                            else:
                                st.warning(f"No geocoding result. {payload.get('error', 'no details')}")
                except requests.RequestException as exc:
                    st.error(f"Could not reach geocode API ({api_base_url}): {exc}")
        st.warning("Coastline tip: keep the pin slightly inland to avoid sparse ocean-side sampling.")
        _loc_for_start = get_scrape_location()
        _pin_source_for_start = str(_loc_for_start.get("source") or "")
        _pin_ready_for_start = _pin_source_for_start not in {"init", ""}
        if (not _pin_ready_for_start) and ("pin_lat" in st.query_params and "pin_lng" in st.query_params):
            _pin_ready_for_start = True
        if not _pin_ready_for_start:
            st.caption("Set a pin first using **Address search** or **Run pin** lat/lng + **Apply typed pin**.")
        run_single = st.button("Start Scraping", type="primary", width="stretch", disabled=not _pin_ready_for_start)

    _pin_widget_scope = str(city_key) if is_city_mode else "custom_pin_mode"
    st.subheader("Run pin (single source for scraping)")
    st.caption(
        "Use the **map** (click background) or **lat/lng** below — both update the same `scrape_location` "
        "sent to the API. **Start Scraping** is in the sidebar."
    )
    loc_ui = get_scrape_location()
    lat_internal_key = f"_run_pin_lat__{_pin_widget_scope}"
    lng_internal_key = f"_run_pin_lng__{_pin_widget_scope}"
    st.session_state.setdefault(lat_internal_key, float(loc_ui["lat"]))
    st.session_state.setdefault(lng_internal_key, float(loc_ui["lng"]))
    source_now = str(loc_ui.get("source") or "")
    # Always mirror authoritative ``scrape_location`` into staging keys (covers legacy_migrated and future sources).
    st.session_state[lat_internal_key] = float(loc_ui["lat"])
    st.session_state[lng_internal_key] = float(loc_ui["lng"])
    # ``st.number_input`` keeps its own session_state per ``key=``; after a map click / geocode / preset the
    # authoritative pin moves but the widgets would still return stale coords and overwrite the pin below.
    if source_now in {"folium_click", "geocode", "city_preset", "init", "map_center_button", "map_click"}:
        st.session_state.pop(f"run_pin_lat_input__{_pin_widget_scope}", None)
        st.session_state.pop(f"run_pin_lng_input__{_pin_widget_scope}", None)

    preview_two_pinned = False
    dual_points_for_map: list[dict[str, float | str]] | None = None
    if preview_two_pinned:
        dual_points_for_map = [
            {
                "slot": "A",
                "lat": float(st.session_state["dual_area_a_lat"]),
                "lng": float(st.session_state["dual_area_a_lng"]),
                "label": str(st.session_state.get("dual_area_a_label") or "Area A"),
            },
            {
                "slot": "B",
                "lat": float(st.session_state["dual_area_b_lat"]),
                "lng": float(st.session_state["dual_area_b_lng"]),
                "label": str(st.session_state.get("dual_area_b_label") or "Area B"),
            },
        ]

    st.subheader("Interactive search map")
    st.caption("Search + Run pin lat/lng are the reliable pin setters. If browser blocks map click wiring, use those controls.")
    if "pin_lat" in st.query_params and "pin_lng" in st.query_params:
        try:
            b_lat = float(str(st.query_params.get("pin_lat", "")).strip())
            b_lng = float(str(st.query_params.get("pin_lng", "")).strip())
            cur = get_scrape_location()
            if abs(b_lat - float(cur["lat"])) > 1e-5 or abs(b_lng - float(cur["lng"])) > 1e-5:
                set_scrape_location(b_lat, b_lng, "Map pin", "map_click")
                sync_legacy_pin_mirror()
        except Exception:
            pass

    _heal_run_pin_widgets_if_stale_default(_pin_widget_scope)
    _auth_pin_before_num_inputs = get_scrape_location()

    pin_done = str(get_scrape_location().get("source") or "") not in ("init", "")
    step1_cls = "step-card done" if pin_done else "step-card active"
    step2_cls = "step-card active" if pin_done else "step-card"
    step3_cls = "step-card"
    cstep1, cstep2, cstep3 = st.columns(3)
    with cstep1:
        st.markdown(f"<div class='{step1_cls}'><b>Step 1</b><br>Set pin (search or lat/lng)</div>", unsafe_allow_html=True)
    with cstep2:
        st.markdown(f"<div class='{step2_cls}'><b>Step 2</b><br>Set radius &amp; label</div>", unsafe_allow_html=True)
    with cstep3:
        st.markdown(f"<div class='{step3_cls}'><b>Step 3</b><br>Download Excel</div>", unsafe_allow_html=True)

    rp1, rp2 = st.columns(2)
    with rp1:
        run_lat = st.number_input(
            "Run pin latitude",
            value=float(st.session_state[lat_internal_key]),
            format="%.6f",
            step=0.0001,
            key=f"run_pin_lat_input__{_pin_widget_scope}",
        )
    with rp2:
        run_lng = st.number_input(
            "Run pin longitude",
            value=float(st.session_state[lng_internal_key]),
            format="%.6f",
            step=0.0001,
            key=f"run_pin_lng_input__{_pin_widget_scope}",
        )
    _rw_lat = float(run_lat)
    _rw_lng = float(run_lng)
    _fin_lat, _fin_lng = _coalesce_run_pin_inputs_vs_authoritative(_rw_lat, _rw_lng, _auth_pin_before_num_inputs)
    if abs(_fin_lat - _rw_lat) > 1e-9 or abs(_fin_lng - _rw_lng) > 1e-9:
        st.session_state[f"run_pin_lat_input__{_pin_widget_scope}"] = _fin_lat
        st.session_state[f"run_pin_lng_input__{_pin_widget_scope}"] = _fin_lng
    st.session_state[lat_internal_key] = _fin_lat
    st.session_state[lng_internal_key] = _fin_lng
    apply_pin_col, _ = st.columns([1, 3])
    with apply_pin_col:
        if st.button("Apply typed pin"):
            cur_now = get_scrape_location()
            set_scrape_location(
                float(_fin_lat),
                float(_fin_lng),
                str(cur_now.get("label") or "Run pin"),
                "manual_form",
            )
            sync_legacy_pin_mirror()
    current_loc = get_scrape_location()
    # Typing new lat/lng must reset map dedupe so the next click (even near the same coords) applies.
    if abs(_fin_lat - float(current_loc["lat"])) > 1e-5 or abs(_fin_lng - float(current_loc["lng"])) > 1e-5:
        st.session_state["runpin_last_click_sig"] = ""
        set_scrape_location(
            _fin_lat,
            _fin_lng,
            str(current_loc.get("label") or "Run pin"),
            "manual_form",
        )
        sync_legacy_pin_mirror()

    # Render preview AFTER pin finalization so it is always wired to effective scrape_location.
    _map_loc = get_scrape_location()
    _gm_key = _get_google_maps_api_key_for_basemap()
    _z = max(3, min(21, int(st.session_state.get("_pin_map_zoom_ui", _default_zoom_for_radius_km(radius_km)))))
    if _gm_key:
        components.html(
            render_google_maps_pin(float(_map_loc["lat"]), float(_map_loc["lng"]), _gm_key, float(radius_km)),
            height=440,
        )
    else:
        _gmap_url = f"https://www.google.com/maps?q={float(_map_loc['lat']):.6f},{float(_map_loc['lng']):.6f}&z={_z}&output=embed"
        st.markdown(
            f'<iframe src="{_gmap_url}" width="100%" height="420" style="border:0;border-radius:10px;" loading="lazy"></iframe>',
            unsafe_allow_html=True,
        )
    _render_google_maps_reference_panel(radius_km)

    st.subheader("Run")
    loc_run = get_scrape_location()
    _run_source = str(loc_run.get("source") or "")
    if _run_source in {"init", ""}:
        st.write("**Run pin:** `Not set yet`")
    elif is_city_mode:
        st.write(
            f"**City (label):** `{UAE_CITY_DISPLAY[city_key]}` · **Run pin sent to API:** "
            f"`{float(loc_run['lat']):.6f}, {float(loc_run['lng']):.6f}` "
            "(adjust numbers above or click the map)"
        )
    else:
        st.write(f"**Run pin:** `{float(loc_run['lat']):.6f}, {float(loc_run['lng']):.6f}`")
    st.write(
        f"Radius: `{radius_km} km` · Profile: `{selected_profile_name}` · "
        f"**Talabat rows → Google Places** backfill when the API has `GOOGLE_MAPS_API_KEY` (profile requests it) · "
        f"Status: `{listing_status_mode}` · New-only: `{new_on_platform_only}` · "
        f"Target label: `{target_area_label.strip() or '—'}`"
    )
    st.caption(
        "Runs usually take a few minutes; large radius, **Complete** profile, or API wall-clock limits can add more time."
    )
    with st.expander("Timeouts, Streamlit → API, and optional Secrets", expanded=False):
        st.markdown(
            """
- **Enqueue read-timeout** (Streamlit Cloud to a raw `http://` IP): prefer **`https://` `API_BASE_URL`** (e.g. Caddy), or raise **`SCRAPER_API_POST_TIMEOUT_SEC`** in Secrets (e.g. `900`).
- **Long Worker jobs:** client poll budget follows **`scrape_wall_clock_sec`**; optional override **`SCRAPER_POLL_TOTAL_TIMEOUT_SEC`** (e.g. `3600`).
- **POST `/scrape` returns quickly** (async enqueue): read cap via env or Secrets **`SCRAPER_API_ENQUEUE_READ_CAP_SEC`** (default ~240s); increase only if you use a **legacy synchronous** `/scrape` that returns full `records` in one response.
"""
        )
    pinned_count = "one"
    use_two_pinned = False
    run_two_pins = False

    loc_fp = get_scrape_location()
    dual_sig = "none"
    dual_sig = "one_pin"

    current_fingerprint = "|".join(
        [
            area_mode,
            city_key if is_city_mode else "custom",
            pinned_count,
            listing_status_mode,
            str(new_on_platform_only),
            selected_profile_name,
            str(include_google_coverage),
            dual_sig,
            target_area_label.strip(),
            f"{float(loc_fp['lat']):.6f}",
            f"{float(loc_fp['lng']):.6f}",
            str(radius_km),
        ]
    )

    def _start_scrape_timer(timer_box):
        stop_event = threading.Event()
        start_ts = time.time()

        def _tick():
            while not stop_event.is_set():
                elapsed = int(time.time() - start_ts)
                mins = elapsed // 60
                secs = elapsed % 60
                try:
                    timer_box.markdown(f"⏱ Scraping · {mins:02d}:{secs:02d} elapsed")
                except Exception:
                    return
                time.sleep(1)

        th = threading.Thread(target=_tick, daemon=True)
        th.start()
        return stop_event, start_ts

    if run_two_pins:
        progress = st.progress(0.0)
        status_box = st.empty()
        timer_box = st.empty()
        stop_event, start_ts = _start_scrape_timer(timer_box)
        a_lab = str(st.session_state.get("dual_area_a_label") or "").strip() or "Area A"
        b_lab = str(st.session_state.get("dual_area_b_label") or "").strip() or "Area B"
        dual_area_df = pd.DataFrame(
            [
                {
                    "lat": float(st.session_state["dual_area_a_lat"]),
                    "lng": float(st.session_state["dual_area_a_lng"]),
                    "label": a_lab,
                    "area_slot": "A",
                },
                {
                    "lat": float(st.session_state["dual_area_b_lat"]),
                    "lng": float(st.session_state["dual_area_b_lng"]),
                    "label": b_lab,
                    "area_slot": "B",
                },
            ]
        )
        profile_cfg = _SCRAPE_PROFILES.get(selected_profile_name, _SCRAPE_PROFILES["Complete"])
        base_payload = {
            "radius_km": float(radius_km),
            "spacing_km": _DEFAULT_SPACING_KM,
            "concurrency": _DEFAULT_CONCURRENCY,
            "scroll_rounds": int(profile_cfg["scroll_rounds"]),
            "scroll_wait_ms": int(profile_cfg["scroll_wait_ms"]),
            "status_filter": "all",
            "just_landed_only": bool(new_on_platform_only),
            "max_sample_points": int(profile_cfg["max_sample_points"]),
            "dedupe_by_vendor_url": _SCRAPE_DEDUPE_BY_VENDOR_URL,
            "high_volume": bool(profile_cfg["high_volume"]),
            "google_places_enrich": bool(profile_cfg["google_places_enrich"]),
            "enrich": bool(profile_cfg.get("enrich", False)),
            "scrape_target_label": None,
            "city": None,
        }
        _wall = profile_cfg.get("scrape_wall_clock_sec")
        if _wall is not None:
            base_payload["scrape_wall_clock_sec"] = int(_wall)
        _dual_poll_budget = _scrape_poll_budget_sec_for_payload(base_payload)
        with st.spinner("Dual-area scrape: running pin A, then pin B..."):
            dual_df, dual_errors = run_dual_area_scrape_via_api(
                api_base_url=api_base_url,
                headers=headers,
                locations_df=dual_area_df,
                base_payload=base_payload,
                timeout_sec=_dual_poll_budget,
            )
        dual_df = ensure_just_landed_columns(dual_df)
        stop_event.set()
        elapsed = int(time.time() - start_ts)
        timer_box.success(f"Done in {elapsed // 60:02d}:{elapsed % 60:02d}")
        progress.progress(1.0)
        previous_success_df = st.session_state.get("last_successful_results_df", pd.DataFrame())
        dual_snap_cur = (
            float(st.session_state["dual_area_a_lat"]),
            float(st.session_state["dual_area_a_lng"]),
            float(st.session_state["dual_area_b_lat"]),
            float(st.session_state["dual_area_b_lng"]),
        )
        prev_dual_snap = st.session_state.get("_last_successful_dual_snapshot")
        dual_snap_ok = _coords_tuple_close(prev_dual_snap, dual_snap_cur)
        if dual_df is not None and not dual_df.empty:
            st.session_state["results_df"] = dual_df
            st.session_state["last_successful_results_df"] = dual_df
            st.session_state["_last_successful_dual_snapshot"] = dual_snap_cur
            st.session_state["last_run_returned_zero"] = False
        elif (
            dual_snap_ok
            and previous_success_df is not None
            and not previous_success_df.empty
        ):
            st.session_state["results_df"] = previous_success_df
            st.session_state["last_run_returned_zero"] = True
        else:
            st.session_state["results_df"] = dual_df if dual_df is not None else pd.DataFrame()
            st.session_state["last_run_returned_zero"] = True
            if (
                dual_df is not None
                and dual_df.empty
                and previous_success_df is not None
                and not previous_success_df.empty
                and not dual_snap_ok
            ):
                st.info(
                    "Dual-area scrape returned 0 rows — **not** showing results from a previous A/B pin pair. "
                    "Adjust pins and run again."
                )
        st.session_state["google_coverage_df"] = pd.DataFrame()
        st.session_state["last_run_done"] = True
        st.session_state["results_fingerprint"] = current_fingerprint
        st.session_state["last_scrape_run_meta"] = {
            "dual_area_mode": True,
            "dual_area_locations": int(len(dual_area_df)),
            "dual_area_errors": int(len(dual_errors)),
        }
        if dual_errors:
            st.warning("Dual-area run finished with some errors: " + "; ".join(dual_errors[:5]))
        if dual_df.empty:
            st.warning("Dual-area run returned no rows.")
        else:
            status_box.info(f"Dual-area scrape completed · rows={len(dual_df):,} · pins={len(dual_area_df)}")

    if run_single:
        progress = st.progress(0.0)
        status_box = st.empty()
        timer_box = st.empty()
        stop_event, start_ts = _start_scrape_timer(timer_box)
        if not include_google_coverage:
            st.session_state["google_coverage_df"] = pd.DataFrame()

        with st.spinner("Cooking up some data magic... 🪄✨"):
            loc_req = get_scrape_location()
            try:
                parse_scrape_pin_or_raise_value_error(loc_req["lat"], loc_req["lng"])
            except ValueError as exc:
                st.error(f"Run pin is invalid — fix lat/lng before scraping. ({exc})")
                st.session_state["results_df"] = pd.DataFrame()
                st.session_state["last_run_done"] = False
            else:
                profile_cfg = _SCRAPE_PROFILES.get(selected_profile_name, _SCRAPE_PROFILES["Complete"])
                payload = {
                    "radius_km": float(radius_km),
                    "spacing_km": _DEFAULT_SPACING_KM,
                    "concurrency": _DEFAULT_CONCURRENCY,
                    "scroll_rounds": int(profile_cfg["scroll_rounds"]),
                    "scroll_wait_ms": int(profile_cfg["scroll_wait_ms"]),
                    "status_filter": listing_status_mode,
                    "just_landed_only": bool(new_on_platform_only),
                    "max_sample_points": int(profile_cfg["max_sample_points"]),
                    "dedupe_by_vendor_url": _SCRAPE_DEDUPE_BY_VENDOR_URL,
                    "high_volume": bool(profile_cfg["high_volume"]),
                    "google_places_enrich": bool(profile_cfg["google_places_enrich"]),
                    "enrich": bool(profile_cfg.get("enrich", False)),
                    "scrape_target_label": target_area_label.strip() or None,
                    "pin_lat": float(loc_req["lat"]),
                    "pin_lng": float(loc_req["lng"]),
                    "client_asserted_pin_lat": float(loc_req["lat"]),
                    "client_asserted_pin_lng": float(loc_req["lng"]),
                }
                _wall_single = profile_cfg.get("scrape_wall_clock_sec")
                if _wall_single is not None:
                    payload["scrape_wall_clock_sec"] = int(_wall_single)
                if is_city_mode:
                    payload["city"] = city_key
                else:
                    payload["city"] = None
                if not target_area_label.strip() and is_city_mode:
                    payload["scrape_target_label"] = UAE_CITY_DISPLAY.get(city_key, city_key)
                _poll_budget_sec = _scrape_poll_budget_sec_for_payload(payload)
                try:
                    fallback_notes: list[str] = []
                    req_headers = dict(headers)
                    def _poll_result(request_id_to_poll: str, total_timeout_sec: float = _SCRAPE_CLIENT_TIMEOUT_SEC) -> dict:
                        deadline = time.time() + float(total_timeout_sec)
                        while time.time() < deadline:
                            poll_resp = requests.get(
                                f"{api_base_url.rstrip('/')}/result/{request_id_to_poll}",
                                headers=req_headers,
                                timeout=30,
                            )
                            if poll_resp.status_code >= 400:
                                raise RuntimeError(_friendly_api_error(poll_resp))
                            poll_data = poll_resp.json() if poll_resp.content else {}
                            poll_status = str((poll_data or {}).get("status") or "").lower()
                            if poll_status == "complete":
                                return poll_data
                            if poll_status == "failed":
                                raise RuntimeError(str((poll_data or {}).get("error") or "Remote scrape job failed"))
                            time.sleep(10)
                        raise RuntimeError("Remote scrape job timed out while waiting for completion.")

                    def _enqueue_read_timeout_sec() -> float:
                        try:
                            raw = str(st.secrets.get("SCRAPER_API_POST_TIMEOUT_SEC", "")).strip()
                            if raw:
                                return max(120.0, min(float(raw), float(_poll_budget_sec)))
                        except Exception:
                            pass
                        return max(120.0, min(float(_SCRAPE_POST_TIMEOUT_SEC), float(_poll_budget_sec)))

                    _read_t = _enqueue_read_timeout_sec()
                    _connect_t = min(60.0, _read_t)

                    def _post_enqueue_read_sec(max_read: float) -> float:
                        cap = float(_SCRAPE_ENQUEUE_READ_CAP_SEC)
                        try:
                            raw = str(st.secrets.get("SCRAPER_API_ENQUEUE_READ_CAP_SEC", "")).strip()
                            if raw:
                                cap = max(30.0, float(raw))
                        except Exception:
                            pass
                        return max(45.0, min(float(max_read), cap))

                    _enqueue_read = _post_enqueue_read_sec(_read_t)

                    def _post_scrape(req_payload: dict, timeout_msg: str) -> tuple[int, dict] | None:
                        attempt_rid = uuid.uuid4().hex
                        post_headers = dict(req_headers)
                        post_headers["X-Request-ID"] = attempt_rid
                        try:
                            enqueue_resp = requests.post(
                                f"{api_base_url.rstrip('/')}/scrape",
                                json=req_payload,
                                headers=post_headers,
                                timeout=(_connect_t, _enqueue_read),
                            )
                            if enqueue_resp.status_code >= 400:
                                return int(enqueue_resp.status_code), {}
                            enqueue_data = enqueue_resp.json() if enqueue_resp.content else {}
                            # Backward compatibility: older API versions return final scrape payload directly from /scrape.
                            if isinstance(enqueue_data, dict) and "records" in enqueue_data:
                                return 200, enqueue_data
                            rid = str(
                                enqueue_data.get("request_id")
                                or enqueue_resp.headers.get("X-Request-ID")
                                or attempt_rid
                            ).strip()
                            if not rid:
                                raise RuntimeError("Missing request_id from /scrape enqueue response")
                            result_data = _poll_result(rid, total_timeout_sec=_poll_budget_sec)
                            return 200, result_data
                        except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectTimeout):
                            status_box.warning(timeout_msg)
                            # POST body may never arrive on slow paths, but the API often already enqueued using X-Request-ID.
                            try:
                                probe = requests.get(
                                    f"{api_base_url.rstrip('/')}/result/{attempt_rid}",
                                    headers=req_headers,
                                    timeout=25,
                                )
                            except requests.exceptions.RequestException:
                                return None
                            if probe.status_code == 404:
                                return None
                            if probe.status_code >= 400:
                                return None
                            try:
                                return (200, _poll_result(attempt_rid, total_timeout_sec=_poll_budget_sec))
                            except Exception:
                                return None

                    scrape_result = _post_scrape(
                        payload,
                        "Primary /scrape hit connect or read timeout (checking whether the job was still enqueued)…",
                    )
                    # Do **not** treat client-side timeout (None) like HTTP 504: auto-retries enqueue extra jobs and
                    # overload a single VM (multiple Playwright runs). Only retry on real 502/504 from the server.
                    if scrape_result is None:
                        raise RuntimeError(
                            "POST /scrape did not return a response in time and no /result job could be recovered. "
                            "This is usually network path (Streamlit Cloud → raw HTTP). Prefer **https://** "
                            "`API_BASE_URL`, or raise **SCRAPER_API_ENQUEUE_READ_CAP_SEC** / **SCRAPER_API_POST_TIMEOUT_SEC** "
                            "in Secrets. Wait for any in-flight scrape on the VM to finish, then try once — avoid "
                            "double-clicking **Start Scraping**."
                        )
                    if int(scrape_result[0]) >= 400:
                        code = int(scrape_result[0])
                        # Hosted gateway timeouts: retry with progressively lighter payloads (same host, one job at a time).
                        if code in (502, 504):
                            fallback_payload = dict(payload)
                            fallback_payload["high_volume"] = False
                            fallback_payload["just_landed_only"] = False
                            fallback_payload["status_filter"] = "all"
                            fallback_payload["max_sample_points"] = min(int(payload["max_sample_points"]), 12)
                            fallback_payload["scroll_rounds"] = 10
                            fallback_notes.append("Fallback 1 used: disabled high-volume, reduced grid sample cap and scroll rounds.")
                            scrape_result = _post_scrape(
                                fallback_payload,
                                "Lighter retry also hit read-timeout. Retrying once with ultra-light settings...",
                            )
                            if scrape_result is None:
                                raise RuntimeError(
                                    "Second /scrape attempt timed out client-side before a response. "
                                    "Use HTTPS for API_BASE_URL or increase enqueue timeouts; avoid stacking scrapes on a small VM."
                                )
                            code = int(scrape_result[0])
                            if code in (502, 504):
                                ultra_payload = dict(fallback_payload)
                                ultra_payload["just_landed_only"] = False
                                ultra_payload["status_filter"] = "all"
                                ultra_payload["max_sample_points"] = min(int(fallback_payload["max_sample_points"]), 4)
                                ultra_payload["scroll_rounds"] = 8
                                ultra_payload["spacing_km"] = 2.5
                                fallback_notes.append("Fallback 2 used: ultra-light profile with very low sample cap and wider spacing.")
                                scrape_result = _post_scrape(
                                    ultra_payload,
                                    "Ultra-light retry also hit read-timeout.",
                                )
                                if scrape_result is None:
                                    raise RuntimeError(
                                        "Third /scrape attempt timed out client-side. Check API host load and network; "
                                        "the VM runs at most one scrape at a time by default — wait for the prior job."
                                    )
                        if int(scrape_result[0]) >= 400:
                            raise RuntimeError(f"HTTP {int(scrape_result[0])} from /scrape enqueue")
                    api_data = scrape_result[1] if scrape_result is not None else {}
                    api_request_id = str(api_data.get("request_id") or "")
                    df = pd.DataFrame(api_data.get("records", []))
                    df = ensure_just_landed_columns(df)
                    if df.empty:
                        status_box.warning("Scrape returned 0 rows. Trying emergency single-point fallback...")
                        emergency_payload = dict(payload)
                        emergency_payload["high_volume"] = False
                        emergency_payload["dedupe_by_vendor_url"] = True
                        emergency_payload["just_landed_only"] = False
                        emergency_payload["status_filter"] = "all"
                        emergency_payload["max_sample_points"] = 1
                        emergency_payload["scroll_rounds"] = 4
                        emergency_payload["scroll_wait_ms"] = min(int(payload["scroll_wait_ms"]), 700)
                        em_rid = uuid.uuid4().hex
                        em_headers = dict(req_headers)
                        em_headers["X-Request-ID"] = em_rid
                        em_data: dict = {}
                        em_resp: requests.Response | None = None
                        try:
                            em_resp = requests.post(
                                f"{api_base_url.rstrip('/')}/scrape",
                                json=emergency_payload,
                                headers=em_headers,
                                timeout=(_connect_t, _enqueue_read),
                            )
                        except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectTimeout):
                            try:
                                probe_em = requests.get(
                                    f"{api_base_url.rstrip('/')}/result/{em_rid}",
                                    headers=req_headers,
                                    timeout=25,
                                )
                            except requests.exceptions.RequestException:
                                probe_em = None
                            if (
                                probe_em is not None
                                and probe_em.status_code != 404
                                and probe_em.status_code < 400
                            ):
                                em_data = _poll_result(em_rid, total_timeout_sec=_poll_budget_sec)
                        if em_resp is not None and em_resp.status_code < 400:
                            em_enqueue_data = em_resp.json() if em_resp.content else {}
                            # Backward compatibility: older API versions return final scrape payload directly.
                            if isinstance(em_enqueue_data, dict) and "records" in em_enqueue_data:
                                em_data = em_enqueue_data
                            else:
                                em_rid2 = str(
                                    em_enqueue_data.get("request_id")
                                    or em_resp.headers.get("X-Request-ID")
                                    or em_rid
                                ).strip()
                                em_data = _poll_result(em_rid2, total_timeout_sec=_poll_budget_sec)
                        if em_data:
                            em_df = pd.DataFrame(em_data.get("records", []))
                            em_df = ensure_just_landed_columns(em_df)
                            if not em_df.empty:
                                df = em_df
                                api_request_id = str(
                                    em_data.get("request_id") or api_request_id
                                )
                                api_data = em_data
                    gdf = pd.DataFrame()
                    st.session_state["last_scrape_city"] = api_data.get("city")
                    meta_run = api_data.get("scrape_run_meta") or {}
                    meta_run.setdefault("request_id", api_request_id)
                    st.session_state["last_scrape_run_meta"] = meta_run
                    elat = meta_run.get("effective_scrape_pin_lat")
                    elng = meta_run.get("effective_scrape_pin_lng")
                    if elat is not None and elng is not None:
                        st.session_state["_last_successful_run_effective_pin"] = (float(elat), float(elng))
                    if include_google_coverage:
                        try:
                            gc_resp = requests.post(
                                f"{api_base_url.rstrip('/')}/google-coverage",
                                json={
                                    "pin_lat": float(loc_req["lat"]),
                                    "pin_lng": float(loc_req["lng"]),
                                    "radius_km": float(radius_km),
                                },
                                headers=req_headers,
                                timeout=80,
                            )
                            if gc_resp.status_code < 400:
                                gc_data = gc_resp.json()
                                gdf = pd.DataFrame(gc_data.get("records", []))
                            else:
                                st.warning(f"Google coverage fetch skipped: {_friendly_api_error(gc_resp)}")
                        except Exception as exc:
                            st.warning(f"Google coverage fetch failed: {exc}")
                    st.session_state["google_coverage_df"] = gdf
                    progress.progress(1.0)
                    status_box.info(f"Remote scrape completed · request_id={api_request_id}")
                    if fallback_notes:
                        st.warning("Primary scrape degraded due to timeouts: " + " ".join(fallback_notes))
                except Exception as exc:
                    st.error(f"Remote API scrape failed: {exc}")
                    df = pd.DataFrame()
                    st.session_state["google_coverage_df"] = pd.DataFrame()

                previous_success_df = st.session_state.get("last_successful_results_df", pd.DataFrame())
                cur_pin = (float(loc_req["lat"]), float(loc_req["lng"]))
                snap_pin = st.session_state.get("_last_successful_snapshot_pin")
                pin_snap_ok = _coords_tuple_close(snap_pin, cur_pin)
                if df is not None and not df.empty:
                    st.session_state["results_df"] = df
                    st.session_state["last_successful_results_df"] = df
                    st.session_state["_last_successful_snapshot_pin"] = cur_pin
                    st.session_state["last_run_returned_zero"] = False
                elif (
                    pin_snap_ok
                    and previous_success_df is not None
                    and not previous_success_df.empty
                ):
                    st.session_state["results_df"] = previous_success_df
                    st.session_state["last_run_returned_zero"] = True
                else:
                    st.session_state["results_df"] = df if df is not None else pd.DataFrame()
                    st.session_state["last_run_returned_zero"] = True
                    if (
                        (df is None or df.empty)
                        and previous_success_df is not None
                        and not previous_success_df.empty
                        and not pin_snap_ok
                    ):
                        st.info(
                            "This scrape returned **0 Talabat rows** — the map pin differs from your last successful run, "
                            "so the app will **not** show another location's table. "
                            "Use Google coverage (if on) or fix the scrape, then try again."
                        )
                st.session_state["last_run_done"] = True
                st.session_state["results_fingerprint"] = current_fingerprint
        stop_event.set()
        elapsed = int(time.time() - start_ts)
        timer_box.success(f"Done in {elapsed // 60:02d}:{elapsed % 60:02d}")

    df = st.session_state.get("results_df", pd.DataFrame())
    if st.session_state.get("last_run_returned_zero", False):
        _zr = st.session_state.get("results_df", pd.DataFrame())
        if _zr is not None and not _zr.empty:
            st.warning(
                "Latest scrape returned **0 Talabat rows** at this pin. The table below is still your **last successful** "
                "scrape at the **same** coordinates (cached until a run returns data)."
            )
        else:
            st.warning(
                "Latest scrape returned **0 Talabat rows** for this pin. No cached table is shown (pin changed or no prior success)."
            )
    if df is None or df.empty:
        gdf_fallback = st.session_state.get("google_coverage_df", pd.DataFrame())
        if isinstance(gdf_fallback, pd.DataFrame) and not gdf_fallback.empty:
            g = gdf_fallback.copy()
            df = pd.DataFrame(
                {
                    "restaurant_name": g.get("name", pd.Series([""] * len(g))),
                    "google_rating": g.get("rating", pd.Series([""] * len(g))),
                    "google_reviews_count": g.get("user_ratings_total", pd.Series([""] * len(g))),
                    "google_place_id": g.get("google_place_id", pd.Series([""] * len(g))),
                    "lat": g.get("lat", pd.Series([None] * len(g))),
                    "lng": g.get("lng", pd.Series([None] * len(g))),
                    "platform": pd.Series(["Google coverage"] * len(g)),
                    "status": pd.Series(["coverage_only"] * len(g)),
                }
            )
    if df is None or df.empty:
        if st.session_state.get("last_run_done"):
            st.warning(
                "**No restaurants extracted.** Try radius **10 km**, status **all** (temporarily), Just Landed **off**, then run again. "
                "If it still returns zero rows, check the API host logs (e.g. VPS: `docker compose … logs` for the `/scrape` request)."
            )
        else:
            loc_empty = get_scrape_location()
            pin_ready = str(loc_empty.get("source") or "").strip() not in ("", "init")
            if pin_ready:
                st.info(
                    "Pin is set. Click **Start Scraping** (sidebar or **Run** above) to load restaurants for this pin."
                )
            else:
                st.info(
                    "Move the pin on the map, set lat/lng, or use **Search address**, then click **Start Scraping**."
                )
        return

    view_df, dropped_cols = compact_output_df(df)
    view_df = polish_dataframe_display_noise(view_df)
    if "platform" not in view_df.columns:
        view_df["platform"] = "Talabat"
    if dropped_cols:
        st.caption(f"Hiding **{len(dropped_cols)}** all-empty columns from table/export (noise reduction).")

    if (
        st.session_state.get("results_fingerprint")
        and st.session_state.get("results_fingerprint") != current_fingerprint
    ):
        st.warning(
            "**Area settings changed** since the table below was built. "
            "Click **Start Scraping** again to refresh."
        )

    meta = st.session_state.get("last_scrape_run_meta") or {}
    if meta and show_advanced:
        with st.expander("Resolved scrape parameters (debug)", expanded=False):
            st.json(meta)

    last_eff = st.session_state.get("_last_successful_run_effective_pin")
    if last_eff and last_eff[0] is not None:
        cur = get_scrape_location()
        if abs(float(cur["lat"]) - float(last_eff[0])) > 1e-5 or abs(float(cur["lng"]) - float(last_eff[1])) > 1e-5:
            st.warning(
                "**Run pin changed** since the scrape that produced the table below. "
                "The API `scrape_run_meta.effective_scrape_pin_*` is for the displayed rows; re-run to align."
            )

    coverage_only = is_google_coverage_only_results(df)
    if coverage_only:
        st.warning(
            f"**Last Talabat scrape returned 0 restaurants.** Showing **{len(df):,} Google Places** near your pin "
            "(coverage API: hotels, chains, POIs). These rows are **not** Talabat listings, vendor URLs, or order data. "
            "Use API host **Logs** for `/scrape` (VPS: `docker compose` logs), try radius **10 km**, status **all**, and confirm the response includes `records`."
        )
    else:
        st.success(
            f"Collected **{len(df):,}** deduped vendor rows (`dedupe_by_vendor_url=true`). "
            "Use **brand_id** for rollups and **branch_sku** for branch-level checks."
        )
        st.caption(
            "Runs use high-volume listing coverage, **vendor pages for many unique restaurants** (API caps), "
            "and Google Places when the API has a Maps key. Tune `RESTAURANT_DETAIL_ENRICH_MAX` / wall clock on the host if runs time out."
        )
    brand_series = df.get("brand_id", df.get("brand_display_name", pd.Series(dtype=str)))
    brand_count = int(pd.Series(brand_series).astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    if brand_count == 0 and "google_place_id" in df.columns:
        gpid = df["google_place_id"].astype(str).str.strip().replace("", pd.NA)
        brand_count = int(gpid.dropna().nunique())
    if brand_count == 0 and "restaurant_name" in df.columns:
        brand_count = int(df["restaurant_name"].astype(str).str.strip().replace("", pd.NA).dropna().nunique())
    rating_series = _pick_rating_series(df)
    avg_rating = float(rating_series.dropna().mean()) if not rating_series.dropna().empty else 0.0
    m1, m2, m3 = st.columns(3)
    m1.metric("Total rows", int(len(df)))
    m2.metric("Unique places (Google)" if coverage_only else "Unique brands", brand_count)
    m3.metric("Avg rating", f"{avg_rating:.2f}" if avg_rating > 0 else "—")
    sanity = build_sanity_check_report(df, float(radius_km))
    with st.expander("Sanity check (rows, brands, cuisine, legal/contact)", expanded=True):
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Rows", int(sanity["rows_total"]))
        s2.metric("Unique brands", int(sanity["unique_brands"]))
        s3.metric("Contact coverage", f'{float(sanity["contact_coverage_pct"]):.1f}%')
        s4.metric("Legal name coverage", f'{float(sanity["legal_name_coverage_pct"]):.1f}%')
        st.caption(
            f'Unique-brand ratio: `{float(sanity["unique_brand_ratio"]):.3f}` · '
            f'Unique cuisines: `{int(sanity["unique_cuisines"])}` · '
            f'Top cuisine share: `{float(sanity["top_cuisine_share_pct"]):.1f}%`'
        )
        if str(sanity.get("status") or "") == "warn":
            st.warning(str(sanity.get("note") or "Sanity checks flagged potential quality issues."))
        else:
            st.success(str(sanity.get("note") or "Sanity checks passed."))

    tab_results, tab_heatmap, tab_outbound = st.tabs(
        ["Results", "Heatmap", "Whitespace"]
    )

    gdf = st.session_state.get("google_coverage_df", pd.DataFrame())

    with tab_results:
        if coverage_only:
            st.info(
                "**Nothing in this table comes from Talabat listings** — it is **Google coverage only** "
                "(nearby Places). Talabat rows always include a **`restaurant_url`** on `talabat.com`. "
                "The last `/scrape` returned **0** listing `records`; use API **Logs** on the host (VPS / `docker compose`) and re-run after fixing scrape errors or filters."
            )
        st.dataframe(view_df, width="stretch", height=460)
        excel_df = build_excel_export_df(view_df)
        loc_x = get_scrape_location()
        excel_summary = {
            "Generated (UTC)": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            "Run pin (lat · lng)": f'{float(loc_x["lat"]):.6f}, {float(loc_x["lng"]):.6f}',
            "Radius (km)": str(radius_km),
            "Scrape profile": selected_profile_name,
            "Rows in export": str(len(excel_df)),
            "Area / target label": target_area_label.strip() or "—",
            "Pin label": str(loc_x.get("label") or "—")[:200],
        }
        rid_x = meta.get("request_id")
        if rid_x:
            excel_summary["API request_id"] = str(rid_x)
        eff_lat = meta.get("effective_scrape_pin_lat")
        eff_lng = meta.get("effective_scrape_pin_lng")
        if eff_lat is not None and eff_lng is not None:
            excel_summary["Effective scrape pin (API)"] = f"{float(eff_lat):.6f}, {float(eff_lng):.6f}"
        excel_bytes = dataframe_to_excel_bytes(excel_df, summary=excel_summary)
        _excel_fn = f"area_intel_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        c1, c2, c3 = st.columns([2, 1, 1])
        c1.download_button(
            "Download Excel",
            data=excel_bytes,
            file_name=_excel_fn,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            help="Spreadsheet includes a **Run summary** sheet plus formatted **Area Intel** data (URLs, filters).",
        )
        c2.download_button(
            "CSV",
            data=view_df.to_csv(index=False).encode("utf-8"),
            file_name="area_intel_results.csv",
            mime="text/csv",
        )
        c3.download_button(
            "JSON",
            data=view_df.to_json(orient="records", force_ascii=False).encode("utf-8"),
            file_name="area_intel_results.json",
            mime="application/json",
        )
        if isinstance(gdf, pd.DataFrame) and not gdf.empty:
            talabat_place_ids = set(df.get("google_place_id", pd.Series(dtype=str)).astype(str).str.strip().str.lower().tolist())
            talabat_place_ids.discard("")
            if "google_place_id" in gdf.columns:
                google_place_ids = gdf["google_place_id"].astype(str).str.strip().str.lower()
                google_only = gdf.loc[~google_place_ids.isin(talabat_place_ids)].copy()
            else:
                google_only = gdf.copy()
            with st.expander("Google-only coverage candidates", expanded=False):
                st.dataframe(google_only, width="stretch", height=240)

    meta_pin_lat = meta.get("effective_scrape_pin_lat")
    meta_pin_lng = meta.get("effective_scrape_pin_lng")
    if meta_pin_lat is not None and meta_pin_lng is not None:
        hm_lat, hm_lng = float(meta_pin_lat), float(meta_pin_lng)
    else:
        # Fallback for older API responses that may not include effective pin fields yet.
        last_eff = st.session_state.get("_last_successful_run_effective_pin")
        if last_eff and len(last_eff) == 2:
            hm_lat, hm_lng = float(last_eff[0]), float(last_eff[1])
        else:
            loc_hm = get_scrape_location()
            hm_lat, hm_lng = float(loc_hm["lat"]), float(loc_hm["lng"])
    with tab_heatmap:
        st.caption(f"Heatmap center uses the effective scrape pin: `{hm_lat:.6f}, {hm_lng:.6f}`")
        render_heatmap(
            df,
            pin_lat=hm_lat,
            pin_lng=hm_lng,
            radius_km=float(radius_km),
            supply_df=st.session_state.get("supply_overlay_df"),
            google_coverage_df=st.session_state.get("google_coverage_df"),
        )

    with tab_outbound:
        render_outbound_prioritization_dashboard(df)


if __name__ == "__main__":
    main()
