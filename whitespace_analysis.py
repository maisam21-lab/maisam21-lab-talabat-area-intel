"""
KitchenPark Whitespace Analysis
================================
Scrapes Talabat for each KP facility pin and produces an Excel report:
  - Sheet "Matrix"    : brand rows × facility columns (✓ = brand present within radius)
  - Sheet "Raw"       : all vendor records with facility attribution
  - Sheet "Facilities": facility list with scrape stats

Usage:
  python whitespace_analysis.py                    # Live facilities only, 10km radius
  python whitespace_analysis.py --all-facilities   # Include future facilities
  python whitespace_analysis.py --radius 5         # 5km radius
  python whitespace_analysis.py --max-pages 50     # Cap pages per area (faster, less complete)
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from area_page_scraper import scrape_vendors_near_pin, vendor_to_row, UAE_AREA_REGISTRY

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("whitespace_analysis")

# ---------------------------------------------------------------------------
# KitchenPark facility list (lat/lng extracted from Google Maps links)
# ---------------------------------------------------------------------------
FACILITIES: list[dict] = [
    # Dubai — Live
    {"name": "Business Bay 1",  "emirate": "Dubai",     "go_live": "Live",       "lat": 25.1894021, "lng": 55.2892571},
    {"name": "JLT (1)",         "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0802475, "lng": 55.1512831},
    {"name": "Motor City",      "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0469210, "lng": 55.2303106},
    {"name": "Arjan",           "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0645000, "lng": 55.2393000},  # Evershine Gardens approx
    {"name": "DSO 1",           "emirate": "Dubai",     "go_live": "Live",       "lat": 25.1282034, "lng": 55.3922505},
    {"name": "Bur Dubai",       "emirate": "Dubai",     "go_live": "Live",       "lat": 25.2460615, "lng": 55.2759454},
    {"name": "IMPZ (1)",        "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0383700, "lng": 55.1861500},
    {"name": "COL",             "emirate": "Abu Dhabi", "go_live": "Live",       "lat": 24.4989329, "lng": 54.4031167},
    {"name": "Mirdif",          "emirate": "Dubai",     "go_live": "Live",       "lat": 25.2347469, "lng": 55.4310875},
    {"name": "Sufouh",          "emirate": "Dubai",     "go_live": "Live",       "lat": 25.1103251, "lng": 55.1780541},
    {"name": "Deira",           "emirate": "Dubai",     "go_live": "Live",       "lat": 25.2698737, "lng": 55.3323663},
    {"name": "Wafi",            "emirate": "Dubai",     "go_live": "Live",       "lat": 25.2297643, "lng": 55.3189516,
     "area_id": 1329, "area_slug": "oud-metha"},  # Dubai Healthcare City (1264) has 0 vendors; use Oud Metha
    {"name": "Quoz (1)",        "emirate": "Dubai",     "go_live": "Live",       "lat": 25.1403704, "lng": 55.2446225},
    {"name": "Arjan (3) - EK",  "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0656802, "lng": 55.2354685},
    {"name": "Hessa (2) - EK",  "emirate": "Dubai",     "go_live": "Live",       "lat": 25.0831785, "lng": 55.2018668},
    {"name": "Raha (1) - EK",   "emirate": "Dubai",     "go_live": "Live",       "lat": 24.4389739, "lng": 54.5742641},  # Abu Dhabi / Al Raha
    {"name": "Sharjah Centre",  "emirate": "Sharjah",   "go_live": "Live",       "lat": 25.3376961, "lng": 55.4008590},
    {"name": "Muwaileh - EK",   "emirate": "Sharjah",   "go_live": "Live",       "lat": 25.3045405, "lng": 55.4698694},
    # Dubai — Future
    {"name": "DIC",             "emirate": "Dubai",     "go_live": "2027-11-22", "lat": 25.0930000, "lng": 55.1528000},  # Dubai Internet City
    {"name": "Jabal Ali",       "emirate": "Dubai",     "go_live": "2027-08-31", "lat": 24.9903930, "lng": 55.1427240},
    {"name": "Warsan",          "emirate": "Dubai",     "go_live": "2027-07-15", "lat": 25.1651944, "lng": 55.4281111},
    # Abu Dhabi — Future
    {"name": "Al Nahyan",       "emirate": "Abu Dhabi", "go_live": "2026-04-15", "lat": 24.4640062, "lng": 54.3792994},
    {"name": "Shamkha (2)",     "emirate": "Abu Dhabi", "go_live": "2026-11-15", "lat": 24.3293056, "lng": 54.6898333},
    # Al Ain — Future
    {"name": "Jimi",            "emirate": "Al Ain",    "go_live": "2026-05-15", "lat": 24.2604203, "lng": 55.7209785},
    # Sharjah — Future
    {"name": "Falah",           "emirate": "Sharjah",   "go_live": "2026-12-15", "lat": 25.3261944, "lng": 55.4719167},
]

# Extra area IDs for areas not near the default registry
# Key = approx area name, value = (area_id, slug)
# These were discovered via probe / binary search
_EXTRA_AREA_HINTS: dict[str, tuple[int, str]] = {
    # Add here as more area IDs are confirmed:
    # "mirdif": (XXXX, "mirdif"),
    # "dso": (XXXX, "dubai-silicon-oasis"),
}


def _find_best_area(lat: float, lng: float) -> tuple[int, str]:
    """Return (area_id, area_slug) for the nearest registered area."""
    from area_page_scraper import find_nearest_registry_area
    result = find_nearest_registry_area(lat, lng)
    if result is None:
        raise ValueError("Area registry is empty")
    _key, area_id, area_slug, dist_km = result
    if dist_km > 40:
        logger.warning("Nearest area is %.1fkm away from pin (%.5f, %.5f) — results may be incomplete", dist_km, lat, lng)
    else:
        logger.info("Using area_id=%d slug=%s (%.1fkm from pin)", area_id, area_slug, dist_km)
    return area_id, area_slug


def scrape_facility(
    facility: dict,
    *,
    radius_km: float = 10.0,
    max_pages: int | None = None,
    page_delay: float = 0.5,
) -> tuple[list[dict], dict]:
    """Scrape Talabat vendors near one KP facility pin."""
    lat, lng = facility["lat"], facility["lng"]
    name = facility["name"]
    logger.info("--- Scraping: %s (%.5f, %.5f) ---", name, lat, lng)

    if "area_id" in facility and "area_slug" in facility:
        area_id, area_slug = facility["area_id"], facility["area_slug"]
        logger.info("Using facility override area_id=%d slug=%s", area_id, area_slug)
    else:
        area_id, area_slug = _find_best_area(lat, lng)
    vendors, meta = scrape_vendors_near_pin(
        lat, lng, radius_km,
        area_id=area_id,
        area_slug=area_slug,
        max_pages=max_pages,
        page_delay=page_delay,
    )
    logger.info(
        "%s → %d vendors in %.1fkm (area %s, %d total scraped)",
        name, len(vendors), radius_km, area_slug,
        meta.get("vendors_collected", 0),
    )
    return vendors, meta


def build_matrix(
    facility_vendors: dict[str, list[dict]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build brand × facility matrix and raw records DataFrame.

    Returns (matrix_df, raw_df).
    """
    facility_names = list(facility_vendors.keys())

    # ── Raw records ──────────────────────────────────────────────────────────
    raw_rows = []
    for facility_name, vendors in facility_vendors.items():
        for v in vendors:
            row = vendor_to_row(v)
            row["kp_facility"] = facility_name
            raw_rows.append(row)
    raw_df = pd.DataFrame(raw_rows)

    if raw_df.empty:
        return pd.DataFrame(), raw_df

    # ── Brand deduplication ───────────────────────────────────────────────────
    # Use restaurantId as brand key — same chain across branches.
    # Fall back to name if restaurantId is missing.
    def brand_key(row):
        rid = row.get("restaurant_id")
        return int(rid) if rid else f"name:{str(row.get('name') or '').strip().lower()}"

    # Build brand info lookup (restaurant_id → canonical name)
    brand_info: dict = {}
    for facility_name, vendors in facility_vendors.items():
        for v in vendors:
            key = brand_key(vendor_to_row(v))
            if key not in brand_info:
                brand_info[key] = {
                    "brand_key": key,
                    "brand_name": str(v.get("name") or "").strip(),
                    "cuisine": str(v.get("cuisineString") or ""),
                    "restaurant_id": v.get("restaurantId"),
                }

    # ── Matrix: brand × facility ──────────────────────────────────────────────
    # Cell value: number of branches within radius (0 = not present)
    matrix_data: dict[tuple, dict[str, int]] = {}
    for facility_name, vendors in facility_vendors.items():
        for v in vendors:
            row = vendor_to_row(v)
            key = brand_key(row)
            if key not in matrix_data:
                matrix_data[key] = {f: 0 for f in facility_names}
            matrix_data[key][facility_name] = matrix_data[key].get(facility_name, 0) + 1

    # Convert to DataFrame
    matrix_rows = []
    for key, presence in matrix_data.items():
        info = brand_info.get(key, {})
        row = {
            "restaurant_id": info.get("restaurant_id"),
            "brand_name": info.get("brand_name", str(key)),
            "cuisine": info.get("cuisine", ""),
        }
        for facility_name in facility_names:
            row[facility_name] = presence.get(facility_name, 0)
        matrix_rows.append(row)

    matrix_df = pd.DataFrame(matrix_rows)
    if not matrix_df.empty:
        # Sort: brands present in most facilities first
        facility_cols = [c for c in matrix_df.columns if c in facility_names]
        matrix_df["_total_facilities"] = (matrix_df[facility_cols] > 0).sum(axis=1)
        matrix_df = matrix_df.sort_values(["_total_facilities", "brand_name"], ascending=[False, True])
        matrix_df = matrix_df.drop(columns=["_total_facilities"])
        matrix_df = matrix_df.reset_index(drop=True)

    return matrix_df, raw_df


def export_excel(
    matrix_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    facilities: list[dict],
    facility_meta: dict[str, dict],
    output_path: str,
    radius_km: float,
    google_gaps_df=None,
) -> None:
    """Write the three-sheet Excel report."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Facilities sheet written FIRST — guaranteed anchor so the workbook
        #    always has at least one visible sheet even if all other data is empty.
        fac_rows = []
        for f in facilities:
            meta = facility_meta.get(f["name"], {})
            fac_rows.append({
                "Facility": f["name"],
                "Emirate": f["emirate"],
                "Go Live": f["go_live"],
                "Latitude": f["lat"],
                "Longitude": f["lng"],
                "Talabat Area": meta.get("area_slug", ""),
                "Vendors in Radius": meta.get("vendors_in_radius", ""),
                "Total Area Vendors": meta.get("total_vendors_reported", ""),
                "Radius (km)": radius_km,
            })
        pd.DataFrame(fac_rows).to_excel(writer, sheet_name="Facilities", index=False)

        # ── Sheet 1: Matrix ───────────────────────────────────────────────────
        if not matrix_df.empty:
            _kp_extra = [c for c in ("kp_tenant", "kp_facilities", "opportunity") if c in matrix_df.columns]
            _enrich_extra = [c for c in ("contact_phone", "phone_type", "legal_name", "google_address", "google_maps_link", "data_source") if c in matrix_df.columns]
            _fixed = ["restaurant_id", "brand_name", "cuisine"] + _kp_extra + _enrich_extra
            facility_cols = [c for c in matrix_df.columns if c not in _fixed]

            # Replace counts with ✓ / blank for readability (keep counts as tooltips via cell value)
            display_df = matrix_df.copy()
            for col in facility_cols:
                display_df[col] = pd.to_numeric(display_df[col], errors="coerce").fillna(0).apply(lambda x: "✓" if x > 0 else "")

            display_df.to_excel(writer, sheet_name="Matrix", index=False)

            ws = writer.sheets["Matrix"]
            # Header row formatting
            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            check_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            check_font = Font(color="375623", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == "✓":
                        cell.fill = check_fill
                        cell.font = check_font
                        cell.alignment = Alignment(horizontal="center")

            # Column widths
            ws.column_dimensions["A"].width = 10   # restaurant_id
            ws.column_dimensions["B"].width = 30   # brand_name
            ws.column_dimensions["C"].width = 22   # cuisine
            # KP extra columns
            from openpyxl.styles import PatternFill as _PF, Font as _F
            _kp_green_fill = _PF(start_color="1B3A26", end_color="1B3A26", fill_type="solid")
            _kp_opp_fill   = _PF(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for i, col_name in enumerate(_kp_extra, start=4):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = 18
                ws.cell(row=1, column=i).fill = _kp_green_fill
                ws.cell(row=1, column=i).font = _F(color="FFFFFF", bold=True)
            # Opportunity column: highlight cells
            if "opportunity" in _kp_extra:
                opp_col_idx = _fixed.index("opportunity") + 1
                for row in ws.iter_rows(min_row=2, min_col=opp_col_idx, max_col=opp_col_idx):
                    for cell in row:
                        if cell.value and "Opportunity" in str(cell.value):
                            cell.fill = _kp_opp_fill
            # Widths for enrichment columns
            _enrich_widths = {"contact_phone": 20, "phone_type": 14, "legal_name": 28, "google_address": 35, "google_maps_link": 14, "data_source": 18}
            for i, col_name in enumerate(_enrich_extra, start=4 + len(_kp_extra)):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = _enrich_widths.get(col_name, 18)
            for i, col in enumerate(facility_cols, start=4 + len(_kp_extra) + len(_enrich_extra)):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = 14
            ws.freeze_panes = ws.cell(row=2, column=4 + len(_kp_extra) + len(_enrich_extra)).coordinate

        # ── Sheet 2: KP Whitespace (opportunity brands) ──────────────────────
        if not matrix_df.empty and "opportunity" in matrix_df.columns:
            from openpyxl.styles import PatternFill, Font, Alignment
            opp_df = matrix_df[matrix_df["opportunity"] == "⭐ Opportunity"].copy()
            opp_cols = ["brand_name", "cuisine", "opportunity", "kp_facilities"] + \
                       [c for c in matrix_df.columns if c not in
                        ("restaurant_id", "brand_name", "cuisine", "kp_tenant", "kp_facilities", "opportunity")]
            opp_cols = [c for c in opp_cols if c in opp_df.columns]
            opp_display = opp_df[opp_cols].copy()
            # Show branch counts (not ✓) in opportunity sheet so it's more informative
            opp_display.to_excel(writer, sheet_name="KP Whitespace", index=False)
            ws2 = writer.sheets["KP Whitespace"]
            opp_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            head_fill = PatternFill(start_color="1B3A26", end_color="1B3A26", fill_type="solid")
            head_font = Font(color="FFFFFF", bold=True)
            for cell in ws2[1]:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == "⭐ Opportunity":
                        cell.fill = opp_fill
            ws2.column_dimensions["A"].width = 30
            ws2.column_dimensions["B"].width = 22
            ws2.freeze_panes = "A2"

        # ── Sheet 3: Google Gaps ─────────────────────────────────────────────
        if google_gaps_df is not None and not google_gaps_df.empty:
            google_gaps_df.to_excel(writer, sheet_name="Google Gaps", index=False)

        # ── Sheet 4: Raw Records ─────────────────────────────────────────────
        if not raw_df.empty:
            raw_df.to_excel(writer, sheet_name="Raw Records", index=False)

        # Guard: openpyxl raises "At least one sheet must be visible" if every
        # sheet in the workbook is hidden. Ensure Facilities (written above) is
        # always visible, and remove any stale default empty "Sheet" entry.
        wb = writer.book
        for ws in list(wb.worksheets):
            if ws.title == "Sheet" and ws.max_row == 1 and ws.max_column == 1:
                wb.remove(ws)
        if wb.worksheets:
            wb.worksheets[0].sheet_state = "visible"

    logger.info("Saved: %s", output_path)


def _checkpoint_path(output: str) -> str:
    p = Path(output)
    return str(p.with_name(p.stem + "_checkpoint.json"))


def _save_checkpoint(
    checkpoint_file: str,
    facility_vendors: dict,
    facility_meta: dict,
    radius_km: float,
) -> None:
    data = {
        "radius_km": radius_km,
        "saved_at": datetime.now().isoformat(),
        "facilities": {
            name: {"vendors": vendors, "meta": facility_meta.get(name, {})}
            for name, vendors in facility_vendors.items()
        },
    }
    with open(checkpoint_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    logger.info("Checkpoint saved: %s (%d facilities)", checkpoint_file, len(facility_vendors))


def _load_checkpoint(checkpoint_file: str) -> tuple[dict, dict, float]:
    with open(checkpoint_file, encoding="utf-8") as f:
        data = json.load(f)
    facility_vendors = {name: entry["vendors"] for name, entry in data["facilities"].items()}
    facility_meta = {name: entry["meta"] for name, entry in data["facilities"].items()}
    radius_km = data.get("radius_km", 10.0)
    logger.info("Loaded checkpoint: %s (%d facilities, radius=%.1fkm)", checkpoint_file, len(facility_vendors), radius_km)
    return facility_vendors, facility_meta, radius_km


def main() -> None:
    parser = argparse.ArgumentParser(description="KitchenPark Talabat whitespace analysis")
    parser.add_argument("--all-facilities", action="store_true", help="Include future (non-Live) facilities")
    parser.add_argument("--radius", type=float, default=10.0, help="Search radius in km (default 10)")
    parser.add_argument("--max-pages", type=int, default=None, help="Cap pages per area scrape (None = all)")
    parser.add_argument("--page-delay", type=float, default=0.5, help="Delay between page fetches (seconds)")
    parser.add_argument("--output", type=str, default=None, help="Output Excel path")
    parser.add_argument("--facility", type=str, default=None, help="Run for one facility only (name match)")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint (skip already-done facilities)")
    parser.add_argument("--export-only", action="store_true", help="Export Excel from checkpoint without scraping")
    args = parser.parse_args()

    # Filter facilities
    facilities = FACILITIES
    if not args.all_facilities:
        facilities = [f for f in facilities if f["go_live"] == "Live"]
        logger.info("Running Live facilities only (%d). Use --all-facilities for all.", len(facilities))
    if args.facility:
        q = args.facility.lower()
        facilities = [f for f in facilities if q in f["name"].lower()]
        if not facilities:
            logger.error("No facility matched %r", args.facility)
            sys.exit(1)

    logger.info("Facilities to scrape: %d", len(facilities))
    logger.info("Radius: %.1f km | Max pages: %s", args.radius, args.max_pages or "all")

    # Output path
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output = args.output or f"kp_whitespace_{timestamp}.xlsx"
    ckpt_file = _checkpoint_path(output)

    # ── Load checkpoint if resuming or export-only ────────────────────────────
    facility_vendors: dict[str, list[dict]] = {}
    facility_meta: dict[str, dict] = {}

    if (args.resume or args.export_only) and Path(ckpt_file).exists():
        facility_vendors, facility_meta, _ = _load_checkpoint(ckpt_file)
        logger.info("Resuming with %d already-done facilities: %s", len(facility_vendors), list(facility_vendors.keys()))
    elif args.export_only:
        logger.error("--export-only requires a checkpoint file at %s", ckpt_file)
        sys.exit(1)

    if args.export_only:
        # Just export from checkpoint, no scraping
        done_facilities = [f for f in facilities if f["name"] in facility_vendors]
        matrix_df, raw_df = build_matrix(facility_vendors)
        export_excel(matrix_df, raw_df, done_facilities, facility_meta, output, radius_km=args.radius)
        print(f"\nExported {len(done_facilities)} facilities from checkpoint -> {output}")
        return

    # ── Run scrapes ───────────────────────────────────────────────────────────
    already_done = set(facility_vendors.keys())
    todo = [f for f in facilities if f["name"] not in already_done]
    if already_done:
        logger.info("Skipping %d already-done: %s", len(already_done), sorted(already_done))
    logger.info("Scraping %d facilities...", len(todo))

    import time as _time
    for i, facility in enumerate(todo, 1):
        logger.info("[%d/%d] %s", i, len(todo), facility["name"])
        try:
            vendors, meta = scrape_facility(
                facility,
                radius_km=args.radius,
                max_pages=args.max_pages,
                page_delay=args.page_delay,
            )
            facility_vendors[facility["name"]] = vendors
            facility_meta[facility["name"]] = meta
        except Exception as exc:
            logger.error("Failed to scrape %s: %s", facility["name"], exc)
            facility_vendors[facility["name"]] = []
            facility_meta[facility["name"]] = {"error": str(exc)}

        # Save checkpoint after every facility
        _save_checkpoint(ckpt_file, facility_vendors, facility_meta, args.radius)

        if i < len(todo):
            _time.sleep(15)

    # ── Build matrix ──────────────────────────────────────────────────────────
    logger.info("Building brand x facility matrix...")
    all_done_facilities = [f for f in facilities if f["name"] in facility_vendors]
    matrix_df, raw_df = build_matrix(facility_vendors)
    logger.info("Matrix: %d unique brands across %d facilities", len(matrix_df), len(all_done_facilities))

    # ── Export ────────────────────────────────────────────────────────────────
    export_excel(matrix_df, raw_df, all_done_facilities, facility_meta, output, radius_km=args.radius)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"  KitchenPark Whitespace Analysis Complete")
    print(f"  Facilities scraped : {len(all_done_facilities)}")
    print(f"  Unique brands found: {len(matrix_df)}")
    print(f"  Total vendor rows  : {len(raw_df)}")
    print(f"  Output             : {output}")
    print("=" * 60)

    if not matrix_df.empty:
        facility_cols = [c for c in matrix_df.columns if c not in ("restaurant_id", "brand_name", "cuisine")]
        present_counts = (matrix_df[facility_cols] > 0).sum(axis=1)
        gaps = matrix_df[present_counts <= 1]
        print(f"\n  Potential expansion targets (in 0-1 facility): {len(gaps)} brands")
        print(f"  Brands in ALL facilities: {(present_counts == len(facility_cols)).sum()}")
        print()


if __name__ == "__main__":
    main()
